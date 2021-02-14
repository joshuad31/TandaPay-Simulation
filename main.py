import ntpath
import os
from datetime import datetime
import random
import sys
from functools import partial

import qdarkstyle
from PySide2.QtWidgets import QMainWindow, QApplication, QFileDialog
from openpyxl import load_workbook

from settings import RESULT_DIR
from ui.ui_tps import Ui_TandaPaySimulationWindow
from utils.common import get_config, update_config_file
from utils.logger import logger


class TandaPaySimulationApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self.ui = Ui_TandaPaySimulationWindow()
        self.ui.setupUi(self)
        self.conf = get_config()
        self.wb = {}
        self.sh = {}
        self.excel_files = {}
        for k in {"system", "user"}:
            getattr(self.ui, f"{k}_database").setText(self.conf['database'][k])
            getattr(self.ui, f"btn_{k}_database").released.connect(partial(self._on_btn_database, k))
            self._init_sheet(k)
        self.ev = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0.]
        self.pv = [0, 0, 0, 0, 0, 0]
        self.rows = [0, 0, 0]
        for i in range(9):
            getattr(self.ui, f"ev_{i}").textChanged.connect(partial(self._on_value_changed, 'ev', i))
            self._on_value_changed('ev', i, int(getattr(self.ui, f"ev_{i}").value()))
        for i in range(6):
            getattr(self.ui, f"pv_{i}").textChanged.connect(partial(self._on_value_changed, 'pv', i))
            self._on_value_changed('pv', i, int(getattr(self.ui, f"pv_{i}").value()))

        self.start_iter = 0
        self.counter = 0
        self.current_period_list = []

        self.sy_rec_p = [None, ] * 21
        self.sy_rec_f = [None, ] * 21
        self.sy_rec_r = [None, ] * 21

        self.ui.btn_exit.released.connect(self.close)
        self.ui.btn_start.released.connect(self.btn_start)
        self.ui.btn_clear.released.connect(self.btn_clear)

    def _on_btn_database(self, db_type: str):
        db_file, _ = QFileDialog.getOpenFileName(
            self, f"Select {db_type.capitalize()} Database File",
            os.path.dirname(self.conf['database'][db_type]), "Excel files (*.xlsx)")
        if db_file:
            self.conf['database'][db_type] = db_file
            update_config_file(self.conf)
            getattr(self.ui, f"{db_type}_database").setText(db_file)
            self._init_sheet(db_type)

    def _init_sheet(self, db_type):
        db_file = self.conf['database'][db_type]
        self.wb[db_type] = load_workbook(db_file)
        self.sh[db_type] = self.wb[db_type].active
        self.excel_files[db_type] = os.path.join(RESULT_DIR, ntpath.basename(self.conf['database'][db_type]))
        if db_type == 'user':
            for row in self.sh['user']['A2:N200']:
                for cell in row:
                    cell.value = None
        elif db_type == 'system':
            for row in self.sh['system']['C2:U37']:
                for cell in row:
                    cell.value = None
        self.save_to_excel(db_type)

    def save_to_excel(self, db_type):
        self.wb[db_type].save(self.excel_files[db_type])

    def _on_value_changed(self, v_type, index, value):
        # All PV values and EV3 ~ EV6 are percentage values.
        ratio = .01 if (v_type == 'pv' or 2 <= index <= 5) else 1
        getattr(self, v_type)[index] = float(value) * ratio
        if v_type == 'ev' and index in {0, 1}:
            self.ev[9] = self.ev[1] * 0.025 * self.ev[0]
        if v_type == 'ev' and index == 0:
            self.ev[0] = int(self.ev[0])

    def _checksum(self, syfunc: int, period: int, line: int):
        c_count = 0
        c_value = 0
        last_checked = 0
        for i in range(self.ev[0]):
            c_us_rec3_val = self.sh['user'].cell(i + 2, 4)
            c_us_rec8_val = self.sh['user'].cell(i + 2, 9)
            if c_us_rec3_val.value == 0 or c_us_rec8_val.value == 'defected':
                continue
            c_us_rec4_val = self.sh['user'].cell(i + 2, 5)
            if c_us_rec3_val.value != last_checked:
                for _i in range(self.ev[0]):
                    c_ur3_sub = self.sh['user'].cell(_i + 2, 4)
                    c_ur8_sub = self.sh['user'].cell(_i + 2, 9)
                    if c_ur3_sub.value == 0 or c_ur8_sub.value == 'defected':
                        continue
                    c_ur4_sub = self.sh['user'].cell(_i + 2, 5)
                    if c_ur3_sub.value == c_us_rec3_val.value:
                        c_count += 1
                        c_value += c_ur4_sub.value
                if c_value % c_count != 0 or c_count != c_us_rec4_val.value:
                    logger.debug(f'______________ Period {period} -> Line {line}')
                    msg = f'SyFunc {syfunc} _checksum failed: c_value % c_count = {c_value % c_count} - ' \
                          f'supposed to be 0.\nc_UsRec3_val:{c_us_rec3_val.value}'
                    logger.error(msg)
                last_checked = c_us_rec3_val.value
                c_count = 0
                c_value = 0

    def _checksum_sr1(self, _sy_rec1_val: int, syfunc: int, period: int, line: int):
        counter = len([i for i in range(self.ev[0]) if self.sh['user'].cell(i + 2, 4).value == 0])
        if self.ev[0] - _sy_rec1_val != counter:
            logger.debug(f'______________ Period {period} -> Line {line}')
            msg = f'SyFunc {syfunc} _checksum_sr1 failed: counter = {counter} - ' \
                  f'supposed to be {self.ev[0] - _sy_rec1_val}'
            logger.error(msg)

    def get_valid_users(self) -> list:
        """
        Returns list of user indexes (for Excel) where User Record 5 is equal to 'valid'
        """
        return [i + 2 for i in range(self.ev[0]) if self.sh['user'].cell(i + 2, 6).value == 'valid']

    def get_select_users(self, _filter: str, u_rec: int) -> list:
        """
        Returns list of user indexes (for Excel) where User Record 'u_rec' is equal to '_filter' argument
        """
        return [i + 2 for i in range(self.ev[0]) if self.sh['user'].cell(i + 2, u_rec + 1).value == _filter]

    def assign_variables(self):
        for i in range(3):
            self.rows[i] = self.start_iter * 3 + i + 2
        for i in range(1, 20):
            self.sy_rec_p[i] = self.sh['system'].cell(self.start_iter * 3 + 2, i + 2)
            self.sy_rec_f[i] = self.sh['system'].cell(self.start_iter * 3 + 3, i + 2)
            self.sy_rec_r[i] = self.sh['system'].cell(self.start_iter * 3 + 4, i + 2)

    def btn_start(self, count=10):
        self.setEnabled(False)
        self.start_iter = 0
        self.counter = 0
        self.current_period_list = []
        while self.start_iter < count:
            self.counter = self.counter + 1
            current_period = 'Period Data {}'.format(self.counter)
            self.current_period_list.append(current_period)
            logger.info('Current period is: {}'.format(self.current_period_list[self.start_iter]))
            if self.counter == 1:
                logger.debug(f'EV1: {self.ev[0]}')
                for i in range(self.ev[0]):
                    val1 = self.sh['user'].cell(i + 2, 1)
                    val1.value = 'user{}'.format(i + 1)
                    # assigning UsRec9 into the database
                    val2 = self.sh['user'].cell(i + 2, 10)
                    val2.value = 0
                    # assigning UsRec10 into the database
                    val3 = self.sh['user'].cell(i + 2, 11)
                    val3.value = 0
                    # assigning UsRec11 into the database
                    val4 = self.sh['user'].cell(i + 2, 12)
                    val4.value = self.ev[0]
                    # assigning us_rec12 into the database
                    val5 = self.sh['user'].cell(i + 2, 13)
                    val5.value = 'yes'
                    # assigning UsRec13 into the database
                    val6 = self.sh['user'].cell(i + 2, 14)
                    val6.value = 0
                self.save_to_excel('user')
                logger.debug('Initial values for UsRec variables set!')
            # PAGE 8, 9
            if self.current_period_list[self.start_iter] == 'Period Data 1':
                for i in range(2):
                    val = self.sh['system'].cell(i + 2, 3)
                    val.value = self.ev[0]
                    val = self.sh['system'].cell(i + 2, 4)
                    val.value = self.ev[9] / self.ev[0]
                    for k in range(5, 21):
                        val = self.sh['system'].cell(i + 2, k)
                        val.value = 0 if k != 18 else 'no'
                    val = self.sh['system'].cell(i + 2, 21)
                    val.value = self.ev[9] / self.ev[0]

                for i in range(3, 30):
                    for k in range(3, 22):
                        val = self.sh['system'].cell(i + 2, k)
                        val.value = 0
                self.save_to_excel('system')
                logger.debug('Initial values for SyRec variables set!')

                # Subgroup  # FUNCTION FOR SUBGROUP EXECUTION
                step1_ev1 = self.ev[0]
                step2 = self.ev[0] / 5
                step3 = round(step2 / 2.3333)
                step4 = step3 * 5
                step5 = step1_ev1 - step4
                step6 = step5 / 6
                step7 = round(step6 / 2)
                step8 = step7 * 6
                step9 = step5 - step8
                step10 = step9 / 7
                step11 = int(step10 / 2)
                step12 = step11 * 7
                step13 = step9 - step12
                step14 = int(step13 / 4)
                step15 = step13 % 4
                if step15 == 0:
                    pass
                if step15 == 1:
                    step3 = step3 - 1
                    step7 = step7 + 1
                if step15 == 2:
                    step3 = step3 - 1
                    step11 = step11 + 1
                if step15 == 3:
                    step3 = step3 - 1
                    step14 = step14 + 2

                # subgroup division code END
                # now assigning number to the group
                # condition checking for group == 4
                group_num = 1
                group_mem_count = 0
                temp_val_four = step14 * 4
                four_grp = []
                dep_num = 0
                for i in range(temp_val_four):
                    d = self.sh['user'].cell(i + 2, 3)
                    d.value = 4
                    a = self.sh['user'].cell(i + 2, 5)
                    a.value = 4
                    label = self.sh['user'].cell(i + 2, 2)
                    # label.value = 'D'
                    label.value = group_num
                    four_grp.append(group_num)
                    label_1 = self.sh['user'].cell(i + 2, 4)
                    # label_1.value = 'D'
                    label_1.value = group_num
                    group_mem_count += 1
                    self.sh['user'].cell(i + 2, 8).value = 'dependent'
                    dep_num += 1
                    if group_mem_count == 4:
                        group_num += 1
                        group_mem_count = 0

                # ('D group assigned!')
                # condition checking for group == 5
                temp_val_five = step3 * 5
                for i in range(temp_val_five):
                    d = self.sh['user'].cell(i + temp_val_four + 2, 3)
                    d.value = 5
                    a = self.sh['user'].cell(i + temp_val_four + 2, 5)
                    a.value = 5
                    label = self.sh['user'].cell(i + temp_val_four + 2, 2)
                    # label.value = 'A'
                    label.value = group_num
                    label_1 = self.sh['user'].cell(i + temp_val_four + 2, 4)
                    # label_1.value = 'A'
                    label_1.value = group_num
                    group_mem_count += 1
                    if group_mem_count == 5:
                        group_num += 1
                        group_mem_count = 0

                # condition checking for group == 6
                temp_val_six = step7 * 6
                for i in range(temp_val_six):
                    d = self.sh['user'].cell(i + temp_val_four + temp_val_five + 2, 3)
                    d.value = 6
                    a = self.sh['user'].cell(i + temp_val_four + temp_val_five + 2, 5)
                    a.value = 6
                    label = self.sh['user'].cell(i + temp_val_four + temp_val_five + 2, 2)
                    # label.value = 'B'
                    label.value = group_num
                    label_1 = self.sh['user'].cell(i + temp_val_four + temp_val_five + 2, 4)
                    # label_1.value = 'B'
                    label_1.value = group_num
                    group_mem_count += 1
                    if group_mem_count == 6:
                        group_num += 1
                        group_mem_count = 0

                # condition checking for group == 7
                temp_val_seven = step11 * 7
                for i in range(temp_val_seven):
                    d = self.sh['user'].cell(i + temp_val_four + temp_val_five + temp_val_six + 2, 3)
                    d.value = 7
                    a = self.sh['user'].cell(i + temp_val_four + temp_val_five + temp_val_six + 2, 5)
                    a.value = 7
                    label = self.sh['user'].cell(i + temp_val_four + temp_val_five + temp_val_six + 2, 2)
                    # label.value = 'C'
                    label.value = group_num
                    label_1 = self.sh['user'].cell(i + temp_val_four + temp_val_five + temp_val_six + 2, 4)
                    # label_1.value = 'C'
                    label_1.value = group_num
                    group_mem_count += 1
                    if group_mem_count == 7:
                        group_num += 1
                        group_mem_count = 0
                self.save_to_excel('user')

                checksum = temp_val_four + temp_val_five + temp_val_six + temp_val_seven
                if checksum != self.ev[0]:
                    raise ValueError(f"Initial group checksum failed: checksum:{checksum} != self.ev[0]:{self.ev[0]}")
                logger.debug({"D": temp_val_four, "A": temp_val_five, "B": temp_val_six, "C": temp_val_seven})

                # setting valid to UsRec5
                for i in range(self.ev[0]):
                    valid_value = self.sh['user'].cell(i + 2, 6)
                    valid_value.value = 'valid'
                self.save_to_excel('user')
                logger.debug(
                    f'group of four members: {step14}, five members: {step3}, six members: {step7}, '
                    f'seven members: {step11}, Total group: {step14 * 4 + step3 * 5 + step7 * 6 + step11 * 7})')

                # Assign 'dependent' to equal EV6
                dependent_pct = dep_num / self.ev[0]
                remaining_pct = self.ev[5] - dependent_pct
                if remaining_pct > 0:
                    unassigned_dep = int(remaining_pct * self.ev[0])
                else:  # FIXME: Correct value?
                    unassigned_dep = 1

                rand_dep_user = sorted(random.sample(range(dep_num + 1, self.ev[0] + 1), unassigned_dep))

                # ROLE1
                # Role1_list = ['low-morale', 'unity-role']
                # EV 4 = Percentage of honest defectors
                role_ev4 = int(self.ev[0] * self.ev[3])
                rand_defectors = sorted(random.sample(range(1, self.ev[0]), role_ev4))
                # EV 5 = Percentage of low-morale members
                role_ev5 = round(self.ev[0] * self.ev[4])
                low_morale_list = []

                if self.ev[4] > 0:
                    while True:
                        n = random.randint(1, self.ev[0])
                        if n not in rand_defectors and n not in low_morale_list:
                            low_morale_list.append(n)
                            if len(low_morale_list) == role_ev5 or len(low_morale_list) + len(
                                    rand_defectors) == self.ev[0]:
                                # if len(low_morale_list) == role_ev5:
                                break
                # Remaining members play a unity role
                # unity_role = self.ev[0] - (role_ev4 - role_ev5)
                # ROLE2
                # percentage of members unwilling to act alone
                # role_ev6 = round(self.ev[0] * self.ev[5])
                # assigning UsRec6, ROLE1 values to excel
                assigned_dep = dep_num
                assigned_indep = 0
                for i in range(self.ev[0]):
                    us_rec6_init = self.sh['user'].cell(i + 2, 7)
                    if i + 1 in rand_defectors:
                        us_rec6_init.value = 'defector'
                    elif i + 1 in low_morale_list:
                        us_rec6_init.value = 'low-morale'

                    us_rec2_init = self.sh['user'].cell(i + 2, 3)
                    if us_rec2_init.value != 4 and i + 1 in rand_dep_user:
                        us_rec7_init = self.sh['user'].cell(i + 2, 8)
                        us_rec7_init.value = 'dependent'
                        assigned_dep += 1
                    elif us_rec2_init.value != 4 and i + 1 not in rand_dep_user:
                        us_rec7_init = self.sh['user'].cell(i + 2, 8)
                        us_rec7_init.value = 'independent'
                        assigned_indep += 1
                self.save_to_excel('user')
                for i in range(self.ev[0]):
                    us_rec6_init = self.sh['user'].cell(i + 2, 7)
                    if us_rec6_init.value != 'defector':
                        if us_rec6_init.value != 'low-morale':
                            us_rec6_init.value = 'unity-role'

                if assigned_dep + assigned_indep != self.ev[0]:
                    logger.error(f'Dependent/independent assignment error')

                self.save_to_excel('user')
                logger.debug('Roles Assigned!')

            #################
            # ___UsFunc1___
            #################
            self.assign_variables()
            """"
            Pay Stage 1
            USER DEFECTION FUNCTION
            """

            if self.current_period_list[self.start_iter] == 'Period Data 1':
                defector_count = 0
                current_group_num = 1

                # Setting defector values in each subgroup
                defected_cache = {}
                defected_subt = {}
                low_morale_cache = []
                lm_def = []
                for i in range(self.ev[0]):
                    us_rec1 = self.sh['user'].cell(i + 2, 2)
                    us_rec13 = self.sh['user'].cell(i + 2, 14)

                    us_rec6 = self.sh['user'].cell(i + 2, 7)
                    us_rec7 = self.sh['user'].cell(i + 2, 8)

                    if us_rec6.value == 'defector' or i == self.ev[0] - 1:
                        if current_group_num not in defected_subt:
                            defected_subt[current_group_num] = 0
                        # PATH 1 for dependent
                        if us_rec1.value == current_group_num or i == self.ev[0] - 1:
                            if us_rec6.value == 'defector':
                                defector_count += 1
                                us_rec13.value = defector_count
                                if us_rec7.value == 'dependent':
                                    defected_subt[current_group_num] += 1
                                if us_rec13.value >= self.ev[6] or us_rec7.value == 'independent':
                                    # PATH 2 (Part 1 - assigning to cache for UsRec 6 & 13 incrementation in next
                                    # for loop)
                                    defected_cache[current_group_num] = defector_count
                        if us_rec1.value != current_group_num or i == self.ev[0] - 1:
                            if current_group_num in defected_cache:
                                if defected_cache[current_group_num] < self.ev[6]:
                                    if defected_subt[current_group_num] != 0:
                                        defected_cache[current_group_num] = defector_count
                                        defected_cache[current_group_num] -= defected_subt[current_group_num]
                                        low_morale_cache.append(current_group_num)
                                        lm_def.append(current_group_num)

                            if defector_count < self.ev[6] and current_group_num not in defected_cache and \
                                    defector_count != 0:
                                # PATH 3 (Part 1 - assigning to cache for UsRec 6 & 7 values in next for loop)
                                low_morale_cache.append(current_group_num)
                            if i != self.ev[0] - 1:
                                defector_count = 1
                                current_group_num = us_rec1.value
                                if current_group_num not in defected_subt:
                                    defected_subt[current_group_num] = 0
                                us_rec13.value = defector_count
                                if us_rec7.value == 'independent':
                                    # PATH 2 (Part 1 - assigning to cache for UsRec 6 & 13 incrementation in next
                                    # for loop)
                                    defected_cache[current_group_num] = defector_count
                                elif us_rec7.value == 'dependent':
                                    defected_subt[current_group_num] += 1
                self.save_to_excel('user')

                # PATH 2 & 3 (Part 2)
                for i in range(self.ev[0]):
                    us_rec1 = self.sh['user'].cell(i + 2, 2)
                    us_rec3 = self.sh['user'].cell(i + 2, 4)
                    us_rec4 = self.sh['user'].cell(i + 2, 5)
                    us_rec5 = self.sh['user'].cell(i + 2, 6)
                    us_rec6 = self.sh['user'].cell(i + 2, 7)
                    us_rec7 = self.sh['user'].cell(i + 2, 8)
                    us_rec8 = self.sh['user'].cell(i + 2, 9)
                    us_rec12 = self.sh['user'].cell(i + 2, 13)
                    us_rec13 = self.sh['user'].cell(i + 2, 14)

                    if us_rec1.value in defected_cache:
                        # PATH 2 (Part 2)
                        us_rec13.value = defected_cache[us_rec1.value]
                        # group_mems = us_rec4.value - us_rec13.value
                        if us_rec6.value == 'defector' and us_rec1.value not in lm_def:
                            # Defectors >= ev7
                            us_rec12.value = 'no'
                            us_rec8.value = 'defected'
                            self.sy_rec_p[1].value -= 1
                            self.sy_rec_p[5].value += 1
                            self.sy_rec_p[3].value += 1

                            for _ in range(self.ev[0]):
                                ur4 = self.sh['user'].cell(_ + 2, 5)
                                ur1 = self.sh['user'].cell(_ + 2, 2)
                                ur3 = self.sh['user'].cell(_ + 2, 4)
                                ur2 = self.sh['user'].cell(_ + 2, 3)

                                if ur4.value != 0:
                                    if us_rec3.value == ur3.value:
                                        ur4.value -= 1
                                        # ur4.value = group_mems
                                        if us_rec1.value == ur1.value:
                                            ur2.value -= 1
                            us_rec3.value = 0
                            us_rec4.value = 0
                            us_rec5.value = 'NR'
                            us_rec8.value = 'NR'
                            us_rec12.value = 'NR'
                            self.save_to_excel('user')

                        if us_rec6.value == 'defector' and us_rec1.value in lm_def and us_rec7.value == 'independent':
                            # Defectors < ev7 and independent defectors exist
                            us_rec12.value = 'no'
                            us_rec8.value = 'defected'
                            self.sy_rec_p[1].value -= 1
                            self.sy_rec_p[5].value += 1
                            self.sy_rec_p[3].value += 1

                            for _ in range(self.ev[0]):
                                ur4 = self.sh['user'].cell(_ + 2, 5)
                                ur1 = self.sh['user'].cell(_ + 2, 2)
                                ur3 = self.sh['user'].cell(_ + 2, 4)
                                ur2 = self.sh['user'].cell(_ + 2, 3)
                                if ur4.value != 0:
                                    if us_rec3.value == ur3.value:
                                        ur4.value -= 1
                                        # ur4.value = group_mems
                                        if us_rec1.value == ur1.value:
                                            ur2.value -= 1
                            us_rec3.value = 0
                            us_rec4.value = 0
                            us_rec5.value = 'NR'
                            us_rec8.value = 'NR'
                            us_rec12.value = 'NR'
                            self.save_to_excel('user')

                    if us_rec1.value in low_morale_cache and us_rec7.value == 'dependent':
                        # PATH 3 (Part 2)
                        if us_rec1.value in defected_cache:
                            us_rec13.value = defected_cache[us_rec1.value]
                        else:
                            us_rec13.value = 0
                        if us_rec6.value == 'defector':
                            us_rec6.value = 'low-morale'
                self.save_to_excel('user')
                self.save_to_excel('system')

                self._checksum(1, int(self.counter), 535)

                #################
            # ___UsFunc2___
            #################
            self.assign_variables()
            """"
            Pay Stage 2
            USER DEFECTION FUNCTION
            """

            if self.current_period_list[self.start_iter] != 'Period Data 1':

                slope = (self.pv[3] - self.pv[1]) / (self.pv[2] - self.pv[0])

                sy_rec19_prev = self.sh['system'].cell(self.rows[0] - 3, 21)
                try:
                    a = float(self.sy_rec_p[19].value)
                    b = float(sy_rec19_prev.value)
                    inc_premium = (a / b) - 1
                except Exception as e:
                    logger.exception(e)
                    logger.debug(f'SyRec19.value: {self.sy_rec_p[19].value}')
                    logger.debug(f'sy_rec19_prev.value: {sy_rec19_prev.value}')
                    logger.debug(f'row1: {self.rows[0]}')
                    logger.debug(f'row1-3: {self.rows[0] - 3}')
                    inc_premium = 0

                valid_users = []
                for i in range(self.ev[0]):
                    us_rec5 = self.sh['user'].cell(i + 2, 6)
                    if us_rec5.value == 'valid':
                        valid_users.append(i + 2)

                if inc_premium >= self.pv[0]:
                    # PATH1
                    skip_percent = (slope * inc_premium - slope * self.pv[0]) + self.pv[1]

                    skip_hash = round(self.sy_rec_p[1].value * skip_percent)
                    skip_users = random.sample(valid_users, skip_hash)

                    for i in range(self.ev[0]):
                        index = i + 2
                        us_rec12 = self.sh['user'].cell(index, 13)
                        if index in skip_users:
                            us_rec12.value = 'no'

                if inc_premium < self.pv[0]:
                    try:
                        num = (self.sy_rec_p[19].value / (float(self.ev[0] / self.ev[0])) - 1)
                        if num >= self.pv[4]:
                            skip_hash = round(self.sy_rec_p[1].value * self.pv[5])
                            skip_users = random.sample(valid_users, skip_hash)

                            # rand_skip_users = []
                            # for _ in range(skip_hash):
                            #     n = random.randint(2,self.ev[0])
                            #     while True:
                            #         if n in rand_skip_users:
                            #             n = random.randint(2,self.ev[0])
                            #         elif n not in rand_skip_users:
                            #             rand_skip_users.append(n)
                            #             break
                            for i in skip_users:
                                us_rec12 = self.sh['user'].cell(i, 13)
                                us_rec12.value = 'no'
                            self.save_to_excel('user')
                        if num < self.pv[4]:
                            # PATH3
                            if self.ev[7] == 0:
                                pass
                            if self.ev[7] == 1 or self.ev[7] == 2 or self.ev[7] == 3:
                                self.ev[7] -= 1
                                # valid_users = []
                                # for i in range(self.ev[0]):
                                #     val = self.sh['user'].cell(i+2,6)
                                #     if val.value == 'valid':
                                #         valid_users.append(i+2)
                                rand_sel = random.choice(valid_users)
                                rand_us_rec12 = self.sh['user'].cell(rand_sel, 13)
                                rand_us_rec12.value = 'no'
                    except ZeroDivisionError:
                        pass

                    self.save_to_excel('user')

            #################
            # ___SyFunc3___                        #Validate premium function
            #################
            self.assign_variables()
            """"
            Pay Stage 3
            Validate premium function
            """
            valid_users = self.get_valid_users()

            path_1 = []
            for i in valid_users:
                us_rec1 = self.sh['user'].cell(i, 2)
                us_rec3 = self.sh['user'].cell(i, 4)
                us_rec4 = self.sh['user'].cell(i, 5)
                us_rec5 = self.sh['user'].cell(i, 6)
                us_rec8 = self.sh['user'].cell(i, 9)
                us_rec12 = self.sh['user'].cell(i, 13)

                if us_rec12.value == 'no':
                    us_rec8.value = 'skipped'
                    path_1.append(i)
                    self.sy_rec_p[1].value -= 1
                    self.sy_rec_p[5].value += 1  # potential incorrect copying or adding

                    for _ in range(self.ev[0]):
                        ur4 = self.sh['user'].cell(_ + 2, 5)
                        ur1 = self.sh['user'].cell(_ + 2, 2)
                        ur3 = self.sh['user'].cell(_ + 2, 4)
                        ur2 = self.sh['user'].cell(_ + 2, 3)

                        if ur4.value != 0:
                            if us_rec3.value == ur3.value:
                                ur4.value -= 1
                                if us_rec1.value == ur1.value:
                                    ur2.value -= 1

                    us_rec8.value = "NR"
                    us_rec3.value = 0
                    us_rec4.value = 0
                    us_rec5.value = "NR"
                    us_rec12.value = "NR"
                    self.save_to_excel('user')

                elif us_rec12.value == 'yes':
                    us_rec8.value = 'paid'

            self.save_to_excel('user')
            self.save_to_excel('system')

            self._checksum(3, int(self.counter), 671)
            self._checksum_sr1(self.sy_rec_p[1].value, 3, int(self.counter), 672)

            #################
            # ___SyFunc4___
            #################
            self.assign_variables()
            """"
            Pay Stage 4
            Invalidate subgroup function
            """

            _path = 0
            for i in range(self.ev[0]):
                ur4 = self.sh['user'].cell(i + 2, 5)
                ur8 = self.sh['user'].cell(i + 2, 9)
                ur5 = self.sh['user'].cell(i + 2, 6)
                ur10 = self.sh['user'].cell(i + 2, 11)
                ur11 = self.sh['user'].cell(i + 2, 12)

                if ur4.value == 1 or ur4.value == 2 or ur4.value == 3:
                    if ur8.value == 'paid':
                        # us_rec8 = 'paid-invalid'
                        ur8.value = 'paid-invalid'
                        # UsRec5 = 'invalid'
                        ur5.value = 'invalid'
                        ur10.value = ur11.value
                        self.sy_rec_p[6].value += 1
            self.save_to_excel('user')
            self.save_to_excel('system')

            self.assign_variables()
            if self.current_period_list[self.start_iter] == 'Period Data 1':
                for k in range(1, 20):
                    self.sy_rec_f[k].value = self.sy_rec_p[k].value
                    self.sy_rec_r[k].value = self.sy_rec_p[k].value

                #################
                # ___SyFunc5___
                #################
                self.assign_variables()
                self.sy_rec_f[9].value = self.sy_rec_f[3].value * self.sy_rec_f[19].value
                self.sh['system'].cell(4, 11).value = self.sy_rec_f[9].value
                self.save_to_excel('system')
            self._checksum(4, int(self.counter), 715)
            self.assign_variables()

            if self.current_period_list[self.start_iter] != 'Period Data 1':
                for k in range(1, 20):
                    self.sy_rec_r[k].value = self.sy_rec_p[k].value

            #################
            # __SyFunc6__            #User quit function
            #################
            self.assign_variables()
            """"
            Reorg Stage 1
            """

            for i in range(self.ev[0]):
                _path = 0
                us_rec1 = self.sh['user'].cell(i + 2, 2)
                us_rec2 = self.sh['user'].cell(i + 2, 3)
                us_rec3 = self.sh['user'].cell(i + 2, 4)
                us_rec4 = self.sh['user'].cell(i + 2, 5)
                us_rec5 = self.sh['user'].cell(i + 2, 6)
                us_rec6 = self.sh['user'].cell(i + 2, 7)
                us_rec7 = self.sh['user'].cell(i + 2, 8)
                us_rec8 = self.sh['user'].cell(i + 2, 9)
                us_rec12 = self.sh['user'].cell(i + 2, 13)
                # 1 2 3 2 4
                if us_rec8.value == 'paid-invalid':
                    if us_rec6.value == 'low-morale':
                        # Path 1
                        prob = random.uniform(0, 1)
                        if prob >= self.ev[8]:
                            _path = 3
                        elif prob < self.ev[8]:
                            _path = 2

                        if _path == 3:
                            if us_rec7.value == 'independent' or us_rec2.value >= 2:
                                # path4
                                self.sy_rec_r[8].value += 1
                            else:
                                _path = 2

                        if _path == 2:
                            us_rec8.value = 'quit'
                            self.sy_rec_r[1].value -= 1
                            self.sy_rec_r[7].value += 1
                            for _i in range(self.ev[0]):
                                ur4 = self.sh['user'].cell(_i + 2, 5)
                                ur3 = self.sh['user'].cell(_i + 2, 4)
                                ur2 = self.sh['user'].cell(_i + 2, 3)
                                ur1 = self.sh['user'].cell(_i + 2, 2)

                                if ur4.value != 0:
                                    if ur3.value == us_rec3.value:
                                        ur4.value -= 1
                                        if ur1.value == us_rec1.value:
                                            ur2.value -= 1
                            us_rec8.value = "NR"
                            us_rec3.value = 0
                            us_rec4.value = 0
                            us_rec5.value = "NR"
                            us_rec12.value = "NR"
            self.save_to_excel('user')
            self.save_to_excel('system')

            for i in range(self.ev[0]):
                _path = 0
                us_rec1 = self.sh['user'].cell(i + 2, 2)
                us_rec2 = self.sh['user'].cell(i + 2, 3)
                us_rec3 = self.sh['user'].cell(i + 2, 4)
                us_rec4 = self.sh['user'].cell(i + 2, 5)
                us_rec5 = self.sh['user'].cell(i + 2, 6)
                us_rec6 = self.sh['user'].cell(i + 2, 7)
                us_rec7 = self.sh['user'].cell(i + 2, 8)
                us_rec8 = self.sh['user'].cell(i + 2, 9)
                us_rec12 = self.sh['user'].cell(i + 2, 13)
                # 1 2 3 2 4
                if us_rec8.value == 'paid-invalid':
                    if us_rec6.value != 'low-morale':
                        if us_rec7.value == 'independent' or us_rec2.value >= 2:
                            # path4
                            self.sy_rec_r[8].value += 1
                        else:
                            _path = 2

                        if _path == 2:
                            us_rec8.value = 'quit'
                            self.sy_rec_r[1].value -= 1
                            self.sy_rec_r[7].value += 1
                            for _i in range(self.ev[0]):
                                ur4 = self.sh['user'].cell(_i + 2, 5)
                                ur3 = self.sh['user'].cell(_i + 2, 4)
                                ur2 = self.sh['user'].cell(_i + 2, 3)
                                ur1 = self.sh['user'].cell(_i + 2, 2)

                                if ur4.value != 0:
                                    if ur3.value == us_rec3.value:
                                        ur4.value -= 1
                                        if ur1.value == us_rec1.value:
                                            ur2.value -= 1
                            us_rec8.value = "NR"
                            us_rec3.value = 0
                            us_rec4.value = 0
                            us_rec5.value = "NR"
                            us_rec12.value = "NR"
            self.save_to_excel('user')
            self.save_to_excel('system')

            self._checksum(6, int(self.counter), 824)
            self._checksum_sr1(self.sy_rec_r[1].value, 6, int(self.counter), 825)

            #################
            # ___SyFunc7___
            #################
            self.assign_variables()
            """"
            Reorg Stage 2
            """

            _path = 0
            loop_reset = False
            invalid_loop = 0  # UsRec4 for group absorbing invalid member not reassigned twice
            found_subgrp = 0
            pass_over = ["defected", "skipped", "quit", "NR"]
            for i in range(self.ev[0]):
                ur8 = self.sh['user'].cell(i + 2, 9)
                ur4 = self.sh['user'].cell(i + 2, 5)
                if ur8.value == 'paid-invalid':

                    if ur4.value == 1:
                        base_ur4 = ur4.value
                        invalid_loop += 1

                        # Chnage ur4 values of group absorbing invalid member
                        if invalid_loop == 1:
                            old_ur4 = 0
                            new_ur4 = 0  # Used to set subgroup so not override values in edge cases
                            for _i in range(self.ev[0]):
                                ur4_sub = self.sh['user'].cell(_i + 2, 5)
                                ur3_sub = self.sh['user'].cell(_i + 2, 4)
                                ur5_sub = self.sh['user'].cell(_i + 2, 6)
                                ur8_sub = self.sh['user'].cell(_i + 2, 9)
                                if new_ur4 == 0:
                                    if ur4_sub.value == 6 and ur8_sub.value not in pass_over and \
                                            ur5_sub.value == 'valid':
                                        found_subgrp = ur3_sub.value
                                        ur4_sub.value = 7
                                        old_ur4 = 6
                                        new_ur4 = 7
                                        _path = 1
                                elif ur4_sub.value == old_ur4 and ur8_sub.value not in pass_over and \
                                        ur5_sub.value == 'valid':
                                    if ur3_sub.value == found_subgrp:
                                        ur4_sub.value = new_ur4

                            if _path != 1:
                                for _i in range(self.ev[0]):
                                    ur4_sub = self.sh['user'].cell(_i + 2, 5)
                                    ur3_sub = self.sh['user'].cell(_i + 2, 4)
                                    ur5_sub = self.sh['user'].cell(_i + 2, 6)
                                    ur8_sub = self.sh['user'].cell(_i + 2, 9)
                                    if new_ur4 == 0:
                                        if ur4_sub.value == 5 and ur8_sub.value not in pass_over and \
                                                ur5_sub.value == 'valid':
                                            found_subgrp = ur3_sub.value
                                            ur4_sub.value = 6
                                            old_ur4 = 5
                                            new_ur4 = 6
                                    elif ur4_sub.value == old_ur4 and ur8_sub.value not in pass_over and \
                                            ur5_sub.value == 'valid':
                                        if ur3_sub.value == found_subgrp:
                                            ur4_sub.value = new_ur4
                        else:       # FIXME: correct value?
                            new_ur4 = 0

                        if invalid_loop == base_ur4:
                            loop_reset = True
                        ur4.value = new_ur4
                        ur3 = self.sh['user'].cell(i + 2, 4)
                        ur3.value = found_subgrp  # !!! not referenced
                        ur5 = self.sh['user'].cell(i + 2, 6)
                        ur5.value = 'valid'
                        ur8.value = 'reorg'
                        ur9 = self.sh['user'].cell(i + 2, 10)
                        ur9.value = ur9.value + 1

                        if loop_reset:
                            invalid_loop = 0
                            loop_reset = False
            self.save_to_excel('user')
            self._checksum(7, int(self.counter), 906)

            loop_reset = False
            invalid_loop = 0  # UsRec4 for group absorbing invalid member not reassigned twice
            _path = 0
            found_subgrp = 0
            reorg_cache = {}
            for i in range(self.ev[0]):
                ur8 = self.sh['user'].cell(i + 2, 9)
                ur3 = self.sh['user'].cell(i + 2, 4)
                ur4 = self.sh['user'].cell(i + 2, 5)
                if ur8.value == 'paid-invalid':
                    if ur4.value == 2:
                        base_ur4 = ur4.value
                        invalid_loop += 1

                        # Change ur4 values of group absorbing invalid member
                        if invalid_loop == 1 and ur3.value not in reorg_cache:
                            old_ur4 = 0
                            new_ur4 = 0  # Used to set subgroup so not override values in edge cases
                            for _i in range(self.ev[0]):
                                ur4_sub = self.sh['user'].cell(_i + 2, 5)
                                ur3_sub = self.sh['user'].cell(_i + 2, 4)
                                ur8_sub = self.sh['user'].cell(_i + 2, 9)
                                if new_ur4 == 0:
                                    if ur4_sub.value == 5 and ur8_sub.value not in pass_over:
                                        found_subgrp = ur3_sub.value
                                        ur4_sub.value = 7
                                        old_ur4 = 5
                                        new_ur4 = 7
                                    elif ur4_sub.value == 4 and ur8_sub.value not in pass_over:
                                        found_subgrp = ur3_sub.value
                                        ur4_sub.value = 6
                                        old_ur4 = 4
                                        new_ur4 = 6
                                elif ur4_sub.value == old_ur4 and ur8_sub.value not in pass_over:
                                    if ur3_sub.value == found_subgrp:
                                        ur4_sub.value = new_ur4

                            reorg_cache.update({ur3.value: found_subgrp})
                        else:
                            new_ur4 = 0

                        if invalid_loop == base_ur4:
                            loop_reset = True
                        if ur3.value in reorg_cache:
                            found_subgrp = reorg_cache[ur3.value]
                            loop_reset = True
                        ur4.value = new_ur4
                        ur3.value = found_subgrp
                        ur5 = self.sh['user'].cell(i + 2, 6)
                        ur5.value = 'valid'
                        ur8.value = 'reorg'
                        ur9 = self.sh['user'].cell(i + 2, 10)
                        ur9.value = ur9.value + 1

                        if loop_reset:
                            invalid_loop = 0
                            loop_reset = False
            self.save_to_excel('user')
            self._checksum(7, int(self.counter), 966)

            loop_reset = False
            invalid_loop = 0  # UsRec4 for group absorbing invalid member not reassigned twice
            _path = 0
            found_subgrp = 0
            reorg_cache = {}
            for i in range(self.ev[0]):
                ur8 = self.sh['user'].cell(i + 2, 9)
                ur4 = self.sh['user'].cell(i + 2, 5)

                if ur8.value == 'paid-invalid':
                    ur3 = self.sh['user'].cell(i + 2, 4)

                    if ur4.value == 3:
                        base_ur4 = ur4.value
                        invalid_loop += 1

                        # Change ur4 values of group absorbing invalid member
                        if invalid_loop == 1 and ur3.value not in reorg_cache:
                            grp_found = 0
                            old_ur4 = 0
                            new_ur4 = 0  # Used to set subgroup so not override values in edge cases
                            for _i in range(self.ev[0]):

                                ur4_sub = self.sh['user'].cell(_i + 2, 5)
                                ur3_sub = self.sh['user'].cell(_i + 2, 4)
                                ur5_sub = self.sh['user'].cell(_i + 2, 6)
                                ur8_sub = self.sh['user'].cell(_i + 2, 9)
                                if new_ur4 == 0:
                                    if ur4_sub.value == 3 and ur3_sub.value != ur3.value and \
                                            ur8_sub.value not in pass_over:  # and ur5_sub.value == 'valid':
                                        found_subgrp = ur3_sub.value
                                        ur4_sub.value = 6
                                        old_ur4 = 3
                                        new_ur4 = 6
                                        _path = 2
                                        ur5_sub.value = 'valid'
                                        ur8_sub.value = 'reorg'
                                        ur9_sub = self.sh['user'].cell(_i + 2, 10)
                                        ur9_sub.value = ur9_sub.value + 1
                                        grp_found = ur3_sub.value
                                elif ur4_sub.value == old_ur4 and ur8_sub.value not in pass_over:
                                    if ur3_sub.value == found_subgrp:
                                        ur4_sub.value = new_ur4
                                        ur5_sub.value = 'valid'
                                        ur8_sub.value = 'reorg'
                                        ur9_sub = self.sh['user'].cell(_i + 2, 10)
                                        ur9_sub.value = ur9_sub.value + 1

                            if grp_found != 0:
                                reorg_cache.update({ur3.value: found_subgrp})

                            if grp_found == 0:
                                for _i in range(self.ev[0]):

                                    ur4_sub = self.sh['user'].cell(_i + 2, 5)
                                    ur3_sub = self.sh['user'].cell(_i + 2, 4)
                                    # ur5_sub = self.sh['user'].cell(_i + 2, 6)
                                    ur8_sub = self.sh['user'].cell(_i + 2, 9)
                                    if new_ur4 == 0:
                                        if ur4_sub.value == 4 and ur8_sub.value not in pass_over:
                                            found_subgrp = ur3_sub.value
                                            # old_subgroup = ur3.value
                                            ur4_sub.value = 7
                                            old_ur4 = 4
                                            new_ur4 = 7
                                            _path = 1

                                    elif ur4_sub.value == old_ur4 and ur8_sub.value not in pass_over:
                                        if ur3_sub.value == found_subgrp:
                                            ur4_sub.value = new_ur4
                                reorg_cache.update({ur3.value: found_subgrp})
                        else:
                            new_ur4 = 0

                        if invalid_loop == base_ur4:
                            loop_reset = True

                        if ur3.value in reorg_cache:
                            found_subgrp = reorg_cache[ur3.value]
                            loop_reset = True

                        if _path == 2:
                            ur4.value = new_ur4
                            ur3.value = found_subgrp
                            ur5 = self.sh['user'].cell(i + 2, 6)
                            ur5.value = 'valid'
                            ur8.value = 'reorg'
                            ur9 = self.sh['user'].cell(i + 2, 10)
                            ur9.value = ur9.value + 1

                        if _path == 1:
                            ur4.value = new_ur4
                            ur3 = self.sh['user'].cell(i + 2, 4)
                            ur3.value = found_subgrp
                            ur5 = self.sh['user'].cell(i + 2, 6)
                            ur5.value = 'valid'
                            ur8.value = 'reorg'
                            ur9 = self.sh['user'].cell(i + 2, 10)
                            ur9.value = ur9.value + 1

                        if loop_reset:
                            invalid_loop = 0
                            loop_reset = False

            self.save_to_excel('user')

            self._checksum(7, int(self.counter), 1074)

            #################
            # ___SyFunc8___    Claims / refunds function
            #################
            self.assign_variables()
            """"
            Reorg Stage 4
            """

            prob = round(random.uniform(0, 1), 2)
            if self.ev[2] > prob:
                self.sy_rec_r[16].value = 'yes'
            elif self.ev[2] < prob:
                self.sy_rec_r[16].value = "no"
                self.sy_rec_r[17].value = self.sy_rec_r[2].value
            self.save_to_excel('system')

            #################
            # ___SyFunc8.5___
            #################
            self.assign_variables()
            """"
            Reorg Stage 4.5
            """
            self.sy_rec_r[11].value = self.sy_rec_r[5].value * self.sy_rec_r[19].value
            self.sy_rec_r[13].value = self.sy_rec_r[6].value * self.sy_rec_r[19].value
            self.save_to_excel('system')

            #################
            # ___SyFunc9___
            #################
            self.assign_variables()
            """"
            Reorg Stage 5
            """

            try:
                self.sy_rec_r[2].value = float(self.ev[0]) / self.sy_rec_r[1].value
            except ZeroDivisionError:
                pass

            try:
                self.sy_rec_r[14].value = self.sy_rec_r[9].value + self.sy_rec_r[11].value + self.sy_rec_r[13].value
            except ZeroDivisionError:
                pass

            try:
                self.sy_rec_r[15].value = self.sy_rec_r[14].value / self.sy_rec_r[1].value
            except ZeroDivisionError:
                pass

            for i in range(self.ev[0]):
                us10 = self.sh['user'].cell(i + 2, 11)
                us11 = self.sh['user'].cell(i + 2, 12)
                if us10.value != 0:
                    us11.value = self.sy_rec_r[2].value + self.sy_rec_r[15].value - us10.value
                    us10.value = 0
                else:
                    sr18 = self.sy_rec_r[18].value
                    us11.value = self.sy_rec_r[2].value + self.sy_rec_r[15].value - sr18 if sr18 is not None else 0
            self.sy_rec_r[19].value = self.sy_rec_r[2].value + self.sy_rec_r[15].value
            self.save_to_excel('user')
            self.save_to_excel('system')

            #################
            # ___SyFunc11___
            #################
            self.assign_variables()
            """"
            Reorg Stage 7
            """
            _path = 0
            if self.current_period_list[self.start_iter] != 'Period Data 10':
                total = self.sy_rec_r[3].value + self.sy_rec_r[5].value + self.sy_rec_r[7].value

                if total > 0:
                    _path = 1
                elif total == 0:
                    _path = 2

                if _path == 1:
                    reorg_row = 1
                    new_pay_row = 1
                    if self.start_iter == 0:
                        reorg_row = 4
                        new_pay_row = 5
                    if self.start_iter == 1:
                        reorg_row = 7
                        new_pay_row = 8
                    if self.start_iter == 2:
                        reorg_row = 10
                        new_pay_row = 11
                    if self.start_iter == 3:
                        reorg_row = 13
                        new_pay_row = 14
                    if self.start_iter == 4:
                        reorg_row = 16
                        new_pay_row = 17
                    if self.start_iter == 5:
                        reorg_row = 19
                        new_pay_row = 20
                    if self.start_iter == 6:
                        reorg_row = 22
                        new_pay_row = 23
                    if self.start_iter == 7:
                        reorg_row = 25
                        new_pay_row = 26
                    if self.start_iter == 8:
                        reorg_row = 28
                        new_pay_row = 29
                    if self.start_iter == 9:
                        reorg_row = 31

                    # copying values of previous to current
                    sy_rec_new_p = [None] * 21
                    for k in range(1, 20):
                        sy_rec_r = self.sh['system'].cell(reorg_row, k + 2)
                        sy_rec_new_p[k] = self.sh['system'].cell(new_pay_row, k + 2)
                        sy_rec_new_p[k].value = sy_rec_r.value

                    # Overwriting values in new row
                    sy_rec_new_p[18].value = sy_rec_new_p[17].value
                    for k in {3, 5, 6, 9, 10, 11, 13, 14, 15, 17}:
                        sy_rec_new_p[k].value = 0
                    self.save_to_excel('system')

                    self._checksum_sr1(sy_rec_new_p[1].value, 11, int(self.counter), 1201)
                    self._checksum(11, int(self.counter), 1202)

            # logging to log file
            if self.current_period_list[self.start_iter] == 'Period Data 10' or _path == 2:
                logger.info(f'Run complete, logging simulation results')
                try:
                    percent = (self.sh['system'].cell(3, 5).value / self.ev[0]) * 100
                    inc_premium = round((self.sy_rec_f[19].value / self.sh["system"].cell(2, 21).value) * 100, 2)
                    result_file = os.path.join(RESULT_DIR, f"{datetime.now().strftime('%m_%d_%Y__%H_%M_%S')}.txt")
                    with open(result_file, 'w') as f:
                        lines = [
                            f'{self.ev[0]} is the number of members at the start of the simulation\n',
                            f'{self.sy_rec_r[1].value} is the number of valid members remaining at the end '
                            f'of the simulation\n',
                            f'{round(((self.ev[0] - self.sy_rec_r[1].value) / self.ev[0]) * 100, 2)}% of '
                            f'policyholders left the group by end of simulation\n',
                            f'{round(self.sh["system"].cell(2, 21).value)} was the initial premium members were '
                            f'asked to pay.\n',
                            f'{inc_premium} is the final premium members were asked to pay.\n',
                            f'Premiums increased by {inc_premium}% by end of simulation\n',
                            f'self.SyRec 3 (period 0 finalize) = {self.sh["system"].cell(3, 5).value}\n',
                            f'{self.ev[3] * 100}% of policyholders who were assigned to defect\n',
                            f'{round(percent, 2)}% of policyholders who actually defected\n',
                            f'{(self.pv[4]) * 100}% was the initial collapse threshold set for PV 5\n'
                        ]
                        f.writelines(lines)
                    logger.info(''.join(lines))
                except Exception as e:
                    logger.exception(e)

            self.start_iter += 1
            if self.start_iter == 11:
                logger.info(f'Iteration {self.start_iter} times complete! Please run the entire application again.')
                return
        self.setEnabled(True)

    def btn_clear(self):
        pass


if __name__ == '__main__':

    app = QApplication(sys.argv)
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyside2())
    ex = TandaPaySimulationApp()

    sys._excepthook = sys.excepthook

    def exception_hook(exctype, value, tb):
        logger.error("=========== Crashed!", exc_info=(exctype, value, tb))
        getattr(sys, "_excepthook")(exctype, value, tb)
        ex.on_crashed()
        sys.exit(1)

    sys.excepthook = exception_hook

    logger.info('========== Starting TandaPay Simulation Application ==========')

    ex.show()
    sys.exit(app.exec_())
