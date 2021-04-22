import os
import time
from datetime import datetime
from openpyxl import load_workbook
from utils.tandapay import TandaPaySimulator

# =============================  Test Data  =========================================
ev4_list = [.24, .26, .27, .28, .29]
pv4_list = [.08, .12, .16, .20]
pv5_list = [.77, .83, .89, .95]
ev3_list = [.40, .55, .70]
ev1_list = [60, 68, 70, 85]
ev5_list = [.10, .20]
ev6_list = [.70, 1.00]
ev7_list = [2, 3]
ev2 = 1000
ev8 = 1
ev9 = .33333
pv1 = .10
pv2 = .02
pv3 = .7
pv6 = .03
# ===================================================================================

_cur_dir = os.path.dirname(os.path.realpath(__file__))
os.makedirs(os.path.join(_cur_dir, 'result'), exist_ok=True)


def create_matrix_result(dataset):

    s_time = time.time()
    results = []
    workbook = load_workbook(os.path.join(_cur_dir, 'databases', '3 Matrix Database.xlsx'))
    sh_map = workbook[workbook.sheetnames[0]]
    sh_log = workbook[workbook.sheetnames[1]]

    for i, d in enumerate(dataset):
        print(f"========== Processing {i}")
        ev = d['ev']
        pv = d['pv']
        sim = TandaPaySimulator(ev=ev, pv=pv, matrix=True)
        r = sim.start_simulation()
        collapsed = 1 if r[1] / r[0] < .5 else 0
        results.append({'ev': ev, 'pv': pv, 'c': collapsed, 'result': r})
        sh_map.cell(i + 5, 1).value = i + 1
        sh_map.cell(i + 5, 2 + ev4_list.index(ev[3])).value = collapsed
        sh_map.cell(i + 5, 7 + pv4_list.index(pv[3])).value = collapsed
        sh_map.cell(i + 5, 11 + pv5_list.index(pv[4])).value = collapsed
        sh_map.cell(i + 5, 15 + ev3_list.index(ev[2])).value = collapsed
        sh_map.cell(i + 5, 18 + ev1_list.index(ev[0])).value = collapsed
        sh_map.cell(i + 5, 22 + ev5_list.index(ev[4])).value = collapsed
        sh_map.cell(i + 5, 24 + ev6_list.index(ev[5])).value = collapsed
        sh_map.cell(i + 5, 26 + ev7_list.index(ev[6])).value = collapsed

        sh_log.cell(i + 2, 1).value = i + 1
        for j in range(7):
            # EV3 ~ EV6 are percentage values
            ratio = 100 if 1 < j < 6 else 1
            sh_log.cell(i + 2, 2 + j).value = ev[j] * ratio
        for j in range(6):
            sh_log.cell(i + 2, 9 + j).value = pv[j] * 100
        for j, v in enumerate(r):
            sh_log.cell(i + 2, 15 + j).value = v

    for i, ev4 in enumerate(ev4_list):
        sh_map.cell(2, 2 + i).value = len([r for r in results if r['ev'][3] == ev4])
        sh_map.cell(3, 2 + i).value = len([r for r in results if r['ev'][3] == ev4 and r['c']])
    for i, pv4 in enumerate(pv4_list):
        sh_map.cell(2, 7 + i).value = len([r for r in results if r['pv'][3] == pv4])
        sh_map.cell(3, 7 + i).value = len([r for r in results if r['pv'][3] == pv4 and r['c']])
    for i, pv5 in enumerate(pv5_list):
        sh_map.cell(2, 11 + i).value = len([r for r in results if r['pv'][4] == pv5])
        sh_map.cell(3, 11 + i).value = len([r for r in results if r['pv'][4] == pv5 and r['c']])
    for i, ev3 in enumerate(ev3_list):
        sh_map.cell(2, 15 + i).value = len([r for r in results if r['ev'][2] == ev3])
        sh_map.cell(3, 15 + i).value = len([r for r in results if r['ev'][2] == ev3 and r['c']])
    for i, ev1 in enumerate(ev1_list):
        sh_map.cell(2, 18 + i).value = len([r for r in results if r['ev'][0] == ev1])
        sh_map.cell(3, 18 + i).value = len([r for r in results if r['ev'][0] == ev1 and r['c']])
    for i, ev5 in enumerate(ev5_list):
        sh_map.cell(2, 22 + i).value = len([r for r in results if r['ev'][4] == ev5])
        sh_map.cell(3, 22 + i).value = len([r for r in results if r['ev'][4] == ev5 and r['c']])
    for i, ev6 in enumerate(ev6_list):
        sh_map.cell(2, 24 + i).value = len([r for r in results if r['ev'][5] == ev6])
        sh_map.cell(3, 24 + i).value = len([r for r in results if r['ev'][5] == ev6 and r['c']])
    for i, ev7 in enumerate(ev7_list):
        sh_map.cell(2, 26 + i).value = len([r for r in results if r['ev'][6] == ev7])
        sh_map.cell(3, 26 + i).value = len([r for r in results if r['ev'][6] == ev7 and r['c']])

    for i in range(26):
        sh_map.cell(4, 2 + i).value = round(sh_map.cell(3, 2 + i).value / sh_map.cell(2, 2 + i).value * 100, 2)

    result_file = os.path.join(_cur_dir, 'result', f"Matrix_{datetime.now().strftime('%m_%d_%Y__%H_%M_%S')}.xlsx")
    workbook.save(result_file)
    workbook.close()

    print(f"Saved to {result_file}, elapsed: {time.time() - s_time}")
    return result_file


if __name__ == '__main__':

    _dataset = []

    for e4 in ev4_list:
        for p4 in pv4_list:
            for p5 in pv5_list:
                for e3 in ev3_list:
                    for e1 in ev1_list:
                        for e5 in ev5_list:
                            for e6 in ev6_list:
                                for e7 in ev7_list:
                                    _dataset.append({
                                        'ev': [e1, ev2, e3, e4, e5, e6, e7, ev8, ev9],
                                        'pv': [.1, .02, .7, p4, p5, .03]
                                    })

    create_matrix_result(_dataset)
