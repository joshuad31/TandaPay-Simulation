#!/usr/bin/env python
# coding: utf-8

# In[ ]:
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import random
import sys
import logging 
import traceback
from datetime import datetime
import shutil
import os
import time
import json
from ev_generator import generate_matrix


logging.basicConfig(level=logging.ERROR, filemode='simulation.log', format='%(levelname)s:%(message)s')

prog_start = time.time()

# Windows File System
root_sys = os.getcwd()
root_us = os.getcwd()

# Mac File System

# Debug
# root_sys = r'C:\\Users\\james\\Desktop\\repository\\tandapay\\TandaPay v2\\TandaPay\\'
# root_us = r'C:\\Users\\james\\Desktop\\repository\\tandapay\\TandaPay v2\\TandaPay\\'

# Dev File System
# root_sys = os.getcwd()
# root_us = os.getcwd()

sys_file = "1 System Database.xlsx"
# if "1 System Database.xlsx" not in os.listdir(root_sys):
#     wb = openpyxl.Workbook()
#     wb.save(root_sys)
# root_sys = r"C:\\Tandapay\\"
path_system = root_sys + r'\\' + sys_file
wb_system = load_workbook(path_system)
# to identify the active sheet
sh_system = wb_system.active

us_file = "2 User Database.xlsx"
# if "2 User Database.xlsx" not in os.listdir(root_us):
#     wb = openpyxl.Workbook()
#     wb.save(root_us)
# root_us = r"C:\\Tandapay\\"
path_user = root_us + r'\\' + us_file
wb_user = load_workbook(path_user)
# to identify the active sheet
sh_user = wb_user.active

matrix_file = '3 Matrix Database.xlsx'
path_matrix = root_us + r'\\' + matrix_file
matrix_wb = load_workbook(path_matrix)
matrix_var_sh = matrix_wb['Variable Map']
matrix_sys_log = matrix_wb['System Log']


for row in sh_user['A2:N200']:
  for cell in row:
    cell.value = None

for row in sh_system['C2:U37']:
  for cell in row:
    cell.value = None

wb_user.save(path_user)
wb_system.save(path_system)

class Simulator:
    def __init__(self, ui=True, _matrix=None, edge=False):
        self.ui = ui
        self._matrix = _matrix
        self.edge = edge # Used to test various edge case variables
        self.init_start = time.time()

    	#Initialize Variables
        self.ev = []
        self.pv = []
        if self.ui == True:
            self.initVariable()

            # Build GUI
            self.app = tk.Tk()
            self.app.title("TandaPay")
            self.app.geometry("600x700")
            self.app.resizable(False, True)

            self.showLabels()
            self.showEntry()
            self.clearAction()
            self.showBtns()
            self.app.mainloop()
            self.startAction()
        elif self._matrix:
            prog_start = time.time()
            times = []
            failures = []
            for run, vector in enumerate(self._matrix):
                loop_start = time.time()
                for row in sh_user['A2:N200']:
                    for cell in row:
                        cell.value = None

                for row in sh_system['C2:U37']:
                    for cell in row:
                        cell.value = None

                wb_user.save(path_user)
                wb_system.save(path_system)
                self.run = run+1
                print(f'____ Run Number {self.run} of {len(self._matrix)}____')
                self.ev = vector[:9]
                self.ev.append(0)
                self.pv = vector[9:]
                self.start_ev = self.ev.copy()
                self.start_pv = self.pv.copy()
                try:
                    self.startAction(vector=True, vector_array=vector)
                except Exception:
                    print(traceback.format_exc())
                    msg = {'vector':vector,'error msg':traceback.format_exc()}
                    failures.append(msg)
                times.append(time.time()-loop_start)
                if run %20==0:
                    avg_elapsed = sum(times)/len(times)
                    runs_left = len(self._matrix) - self.run
                    estimated_left = runs_left*avg_elapsed
                    print(f'Elapsed: {time.time()-prog_start}')
                    print(f'Estimated completion time: {estimated_left/60} minutes ({estimated_left/60/60} hours)')
            with open('failed_vectors.json', 'w') as f:
                f.write(json.dumps(failures,indent=4))

    def initVariable(self):
        for i in range(10):
            self.ev.append(0)
        for i in range(6):
            self.pv.append(0)

    def create_worksheet(self):
        if '2 User Database.xlsx' not in os.listdir():
            with open('2 User Database.xlsx') as f:
                f.write()
            
    def showBtns(self):
        startBtn = tk.Button(self.app, text="Start", width=10, height=1, bg='green', fg='white', command=self.startAction)
        startBtn.grid(row=17, column=1, padx=30, sticky='w')
        
        stopBtn = tk.Button(self.app, text="Stop", width=10, height=1, bg='red', fg='white', command=self.stopAction)
        stopBtn.grid(row=17, column=2, padx=30, sticky='w')
        
        clearBtn = tk.Button(self.app, text="Clear", width=10, height=1, command=self.clearAction)
        clearBtn.grid(row=17, column=3, padx=30, sticky='w')
        
        closeBtn = tk.Button(self.app, text="Close", width=10, height=1, command=self.closeAction)
        closeBtn.grid(row=17, column=4, padx=30, sticky='w')

    def keypress1(self, event):
        # if event.char not in '0123456789':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable.")
        #     return

        if event.char in '0123456789':
            self.ev[0] = self.ev[0] * 10 + int(event.char)

        if self.ev[0] > 0 and self.ev[0] < 4:
            messagebox.showinfo("Alert!", "Minimum member in the group is 4")

    def keypress2(self, event):
        # if event.char not in '0123456789':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable.")
        #     return

        if event.char in '0123456789':
            self.ev[1] = self.ev[1] * 10 + int(event.char)
            
    def keypress3(self, event):
        # if event.char not in '0123456789':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable.")
        #     return

        if event.char in '0123456789':
            self.ev[2] = self.ev[2] * 10 + int(event.char)

        if self.ev[2] > 75 and self.ev[2] < 25:
            messagebox.showinfo("Alert!", "The value should be between 25 to 75")

    def keypress4(self, event):
        # if event.char not in '0123456789':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable.")
        #     return

        if event.char in '0123456789':
            self.ev[3] = self.ev[3] * 10 + int(event.char)

        if self.ev[3] > 45 and self.ev[3] < 0:
            messagebox.showinfo("Alert!", "The value should be between 10 to 30")
        
    def keypress5(self, event):
        # if event.char not in '0123456789':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable.")

        if event.char in '0123456789':
            self.ev[4] = self.ev[4] * 10 + int(event.char)

        if self.ev[4] > 30 and self.ev[4] < 10:
            messagebox.showinfo("Alert!", "The value should be between 10 to 30")
            
    def keypress6(self, event):
        # if event.char not in '0123456789':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable.")
        #     return

        if event.char in '0123456789':
            self.ev[5] = self.ev[5] * 10 + int(event.char)

        if self.ev[5] > 80 and self.ev[5] < 20:
            messagebox.showinfo("Alert!", "The value should be between 20 to 80")
            
    def keypress7(self, event):
        # if event.char not in '234':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable. Try entering value 2, 3, 4")
        #     return

        if event.char in '234':
            self.ev[6] = self.ev[6] * 10 + int(event.char)

        if self.ev[6] > 4 and self.ev[6] < 2:
            messagebox.showinfo("Alert!", "Enter value 2, 3, 4 as per your requirement.")
            
    def keypress8(self, event):
        # if event.char not in '0123':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable. Try entering value 0, 1, 2, 3")
        #     return

        if event.char in '0123':
            self.ev[7] = self.ev[7] * 10 + int(event.char)

        if self.ev[7] > 4 and self.ev[7] < 2:
            messagebox.showinfo("Alert!", "Enter value 0, 1, 2, 3 as per your requirement.")
            
    def keypress9(self, event):
        print(event)

    def keypress10(self, event):
        print(event)

    def keypress1pv(self, event):
        
        # if event.char not in '0123456789':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable.")
        #     return
        if event.char in '0123456789':
            self.pv[0] = self.pv[0] * 10 + int(event.char)
        if self.pv[0] > 100 and self.ev[0] < 1:
            messagebox.showinfo("Alert!", "Enter value between 1-100")

    def keypress2pv(self, event):
        # if event.char not in '0123456789':
        #     messagebox.showinfo("Alert!", "Sorry. Only numbers are acceptable.")
        #     return
        if event.char in '0123456789':
            self.pv[1] = self.pv[1] * 10 + int(event.char)
        if self.pv[1] > 100 and self.ev[1] < 1:
            messagebox.showinfo("Alert!", "Enter value between 1-100")

    def keypress3pv(self, event):
        return
    def keypress4pv(self, event):
        return
    def keypress5pv(self, event):
        return
    def keypress6pv(self, event):
        return
    def showLabels(self):
        # labels for EV
        for i in range(9):
            ev_lbl = tk.Label(self.app, text="EV " + str(i + 1) + ":")
            ev_lbl.grid(row=i, column=0, padx=10, pady=10, sticky='w')  
        # labels for PV
        for i in range(6):
            ev_lbl = tk.Label(self.app, text="PV " + str(i + 1) + ":")
            ev_lbl.grid(row=i+10, column=0, padx=10, pady=10, sticky='w')

    def showEntry(self):
        self.entry1 = tk.Entry(self.app, width=20)
        self.entry1.bind("<Key>", self.keypress1)
        self.entry1.grid(row=0, column=1, pady=10, sticky='w')

        self.entry2 = tk.Entry(self.app, width=20)
        self.entry2.bind("<Key>", self.keypress2)
        self.entry2.grid(row=1, column=1, pady=10, sticky='w')

        self.entry3 = tk.Entry(self.app, width=20)
        self.entry3.bind("<Key>", self.keypress3)
        self.entry3.grid(row=2, column=1, pady=10, sticky='w')

        self.entry4 = tk.Entry(self.app, width=20)
        self.entry4.bind("<Key>", self.keypress4)
        self.entry4.grid(row=3, column=1, pady=10, sticky='w')

        self.entry5 = tk.Entry(self.app, width=20)
        self.entry5.bind("<Key>", self.keypress5)
        self.entry5.grid(row=4, column=1, pady=10, sticky='w')

        self.entry6 = tk.Entry(self.app, width=20)
        self.entry6.bind("<Key>", self.keypress6)
        self.entry6.grid(row=5, column=1, pady=10, sticky='w')

        self.entry7 = tk.Entry(self.app, width=20)
        self.entry7.bind("<Key>", self.keypress7)
        self.entry7.grid(row=6, column=1, pady=10, sticky='w')

        self.entry8 = tk.Entry(self.app, width=20)
        self.entry8.bind("<Key>", self.keypress8)
        self.entry8.insert(0, "3")
        self.entry8.grid(row=7, column=1, pady=10, sticky='w')

        self.entry9 = tk.Entry(self.app, width=20)
        self.entry9.bind("<Key>", self.keypress9)
        self.entry9.insert(0, "0.3333")
        self.entry9.grid(row=8, column=1, pady=10, sticky='w')

        #entry values for PV varibles
        self.entry1pv = tk.Entry(self.app, width=20)
        self.entry1pv.bind("<Key>", self.keypress1pv)
        self.entry1pv.grid(row=10, column=1, pady=10, sticky='w')

        self.entry2pv = tk.Entry(self.app, width=20)
        self.entry2pv.bind("<Key>", self.keypress2pv)
        self.entry2pv.grid(row=11, column=1, pady=10, sticky='w')

        self.entry3pv = tk.Entry(self.app, width=20)
        self.entry3pv.bind("<Key>", self.keypress3pv)
        self.entry3pv.grid(row=12, column=1, pady=10, sticky='w')

        self.entry4pv = tk.Entry(self.app, width=20)
        self.entry4pv.bind("<Key>", self.keypress4pv)
        self.entry4pv.grid(row=13, column=1, pady=10, sticky='w')

        self.entry5pv = tk.Entry(self.app, width=20)
        self.entry5pv.bind("<Key>", self.keypress5pv)
        self.entry5pv.grid(row=14, column=1, pady=10, sticky='w')

        self.entry6pv = tk.Entry(self.app, width=20)
        self.entry6pv.bind("<Key>", self.keypress6pv)
        self.entry6pv.grid(row=15, column=1, pady=10, sticky='w')

    def matrix_run(self):
        pass

    #the enitre opertion starts here after the start button is clicked    
    def startAction(self, vector=None, vector_array=None):

        def report(msg, _v):
            with open('checksum_failures.txt', 'a') as f:
                f.write('____________________\n')
                f.write(msg + '\n')
                f.write(', '.join(str(v) for v in _v))
                f.write('\n')

        def _checksum(syfunc: int, period: int, line: int):
            c_count = 0
            c_value = 0
            last_checked = 0
            for i in range(self.ev1):
                c_UsRec3_val = sh_user.cell(i+2,4)
                c_UsRec8_val = sh_user.cell(i+2,9)
                if c_UsRec3_val.value == 0 or c_UsRec8_val.value == 'defected':
                    continue
                c_UsRec4_val = sh_user.cell(i+2,5)
                if c_UsRec3_val.value != last_checked:
                    for _i in range(self.ev1):
                        c_ur3_sub = sh_user.cell(_i+2,4)
                        c_ur8_sub = sh_user.cell(_i+2,9)
                        if c_ur3_sub.value == 0 or c_ur8_sub.value == 'defected':
                                continue
                        c_ur4_sub = sh_user.cell(_i+2,5)
                        if c_ur3_sub.value == c_UsRec3_val.value:
                            c_count += 1
                            c_value += c_ur4_sub.value
                    if c_value%c_count != 0 or c_count != c_UsRec4_val.value:
                        print('______________')
                        print(f'Period {period}')
                        print(f'-> Line {str(line)}')
                        if self._matrix:
                            run_log_index = self.run-1
                            msg = f'Run {run_log_index}: SyFunc {syfunc} _checksum failed: c_value % c_count = {c_value%c_count} - supposed to be 0.\nc_UsRec3_val:{c_UsRec3_val.value}'
                        else:
                            msg = f'SyFunc {syfunc} _checksum failed: c_value % c_count = {c_value%c_count} - supposed to be 0.\nc_UsRec3_val:{c_UsRec3_val.value}'
                        logging.error(msg)
                        if vector_array:
                            report(msg, vector_array)
                        # time.sleep(5)
                        # raise ValueError(f'SyFunc 3 Checksum failed: c_value % c_count = {c_value%c_count} - supposed to be 0.\nc_UsRec3_val:{c_UsRec3_val.value}')
                    last_checked = c_UsRec3_val.value
                    c_count = 0
                    c_value = 0
        
        def _checksum_sr1(_syRec1_val: int, syfunc: int, period: int, line: int):
            self.ev1 = self.ev[0]
            counter = 0
            for i in range(self.ev1):
                c_UsRec_3 = sh_user.cell(i+2, 4)
                if c_UsRec_3.value == 0:
                    counter += 1
            if self.ev1 - _syRec1_val != counter:
                print('______________')
                print(f'Period {period}')
                print(f'-> Line {str(line)}')
                if self._matrix:
                    run_log_index = self.run-1
                    msg = f'Run {run_log_index}: SyFunc {syfunc} _checksum_sr1 failed: counter = {counter} - supposed to be {self.ev1 - _syRec1_val}'
                else:
                    msg = f'SyFunc {syfunc} _checksum_sr1 failed: counter = {counter} - supposed to be {self.ev1 - _syRec1_val}'
                print(msg)
                logging.error(msg)
                if vector_array:
                    report(msg, vector_array)
                

        start_iter=0
        counter = 0
        current_period_list=[]

        def get_valid_users() -> list:
            """
            Returns list of user indexes (for Excel) where User Record 5
            is equal to 'valid'
            """
            valid_users = []
            for i in range(self.ev1):
                UsRec5 = sh_user.cell(i+2,6)
                if UsRec5.value == 'valid':
                    valid_users.append(i+2)
            return valid_users

        def get_select_users(_filter: str, u_rec: int) -> list:
            """
            Returns list of user indexes (for Excel) where User Record 'u_rec'
            is equal to '_filter' argument
            """
            select_users = []
            for i in range(self.ev1):
                UsRec = sh_user.cell(i+2,u_rec+1)
                if UsRec.value == _filter:
                    select_users.append(i+2)
            return select_users
        
        def assign_variables():
            """
            Assignes EV, PV, and System Record values based on current
            Period Data
            """
            self.ev1 = self.ev[0]
            self.ev3 = self.ev[2]
            self.ev4 = self.ev[3]
            self.ev5 = self.ev[4]
            self.ev6 = self.ev[5]
            self.ev7 = self.ev[6]
            self.ev8 = self.ev[7]
            self.ev9 = self.ev[8]
            self.ev10 = self.ev[9]
            self.pv1 = self.pv[0]
            self.pv2 = self.pv[1]
            self.pv3 = self.pv[2]
            self.pv4 = self.pv[3]
            self.pv5 = self.pv[4]
            self.pv6 = self.pv[5]

            get_period = current_period_list.index(current_period_list[start_iter])
            self.row1=1
            self.row2=1
            self.row3=1
            if get_period==0:
                self.row1=2
                self.row2=3
                self.row3=4
            if get_period==1:
                self.row1=5
                self.row2=6
                self.row3=7
            if get_period==2:
                self.row1=8
                self.row2=9
                self.row3=10
            if get_period==3:
                self.row1=11
                self.row2=12
                self.row3=13
            if get_period==4:
                self.row1=14
                self.row2=15
                self.row3=16
            if get_period==5:
                self.row1=17
                self.row2=18
                self.row3=19
            if get_period==6:
                self.row1=20
                self.row2=21
                self.row3=22
            if get_period==7:
                self.row1=23
                self.row2=24
                self.row3=25
            if get_period==8:
                self.row1=26
                self.row2=27
                self.row3=28
            if get_period==9:
                self.row1=29
                self.row2=30
                self.row3=31

            self.SyRec1_p = sh_system.cell(self.row1,3)
            self.SyRec1_f = sh_system.cell(self.row2,3)
            self.SyRec1_r = sh_system.cell(self.row3,3)  

            self.SyRec2_p = sh_system.cell(self.row1,4)
            self.SyRec2_f = sh_system.cell(self.row2,4)
            self.SyRec2_r = sh_system.cell(self.row3,4)

            self.SyRec3_p = sh_system.cell(self.row1,5)
            self.SyRec3_f = sh_system.cell(self.row2,5)
            self.SyRec3_r = sh_system.cell(self.row3,5)

            self.SyRec4_p = sh_system.cell(self.row1,6)
            self.SyRec4_f = sh_system.cell(self.row2,6)
            self.SyRec4_r = sh_system.cell(self.row3,6)

            self.SyRec5_p = sh_system.cell(self.row1,7)
            self.SyRec5_f = sh_system.cell(self.row2,7)
            self.SyRec5_r = sh_system.cell(self.row3,7)

            self.SyRec6_p = sh_system.cell(self.row1,8)
            self.SyRec6_f = sh_system.cell(self.row2,8)
            self.SyRec6_r = sh_system.cell(self.row3,8)

            self.SyRec7_p = sh_system.cell(self.row1,9)
            self.SyRec7_f = sh_system.cell(self.row2,9)
            self.SyRec7_r = sh_system.cell(self.row3,9)

            self.SyRec8_p = sh_system.cell(self.row1,10)
            self.SyRec8_f = sh_system.cell(self.row2,10)
            self.SyRec8_r = sh_system.cell(self.row3,10)

            self.SyRec9_p = sh_system.cell(self.row1,11)
            self.SyRec9_f = sh_system.cell(self.row2,11)
            self.SyRec9_r = sh_system.cell(self.row3,11)

            self.SyRec10_p = sh_system.cell(self.row1,12)
            self.SyRec10_f = sh_system.cell(self.row2,12)
            self.SyRec10_r = sh_system.cell(self.row3,12)

            self.SyRec11_p = sh_system.cell(self.row1,13)
            self.SyRec11_f = sh_system.cell(self.row2,13)
            self.SyRec11_r = sh_system.cell(self.row3,13)

            self.SyRec12_p = sh_system.cell(self.row1,14)
            self.SyRec12_f = sh_system.cell(self.row2,14)
            self.SyRec12_r = sh_system.cell(self.row3,14)

            self.SyRec13_p = sh_system.cell(self.row1,15)
            self.SyRec13_f = sh_system.cell(self.row2,15)
            self.SyRec13_r = sh_system.cell(self.row3,15)

            self.SyRec14_p = sh_system.cell(self.row1,16)
            self.SyRec14_f = sh_system.cell(self.row2,16)
            self.SyRec14_r = sh_system.cell(self.row3,16)

            self.SyRec15_p = sh_system.cell(self.row1,17)
            self.SyRec15_f = sh_system.cell(self.row2,17)
            self.SyRec15_r = sh_system.cell(self.row3,17)

            self.SyRec16_p = sh_system.cell(self.row1,18)
            self.SyRec16_f = sh_system.cell(self.row2,18)
            self.SyRec16_r = sh_system.cell(self.row3,18)

            self.SyRec17_p = sh_system.cell(self.row1,19)
            self.SyRec17_f = sh_system.cell(self.row2,19)
            self.SyRec17_r = sh_system.cell(self.row3,19)

            self.SyRec18_p = sh_system.cell(self.row1,20)
            self.SyRec18_f = sh_system.cell(self.row2,20)
            self.SyRec18_r = sh_system.cell(self.row3,20)

            self.SyRec19_p = sh_system.cell(self.row1,21)
            self.SyRec19_f = sh_system.cell(self.row2,21)
            self.SyRec19_r = sh_system.cell(self.row3,21)

        while start_iter<10:
            counter=counter+1
            current_period = 'Period Data {}'.format(counter)
            current_period_list.append(current_period)
            # print('The Current period is: {}'.format(current_period_list[start_iter]))

            if counter == 1:
                if not vector:
                    self.ev[0] = int(self.entry1.get())
                    self.ev1 = int(self.ev[0])
                    print(self.ev1)
                    
                    self.ev[1] = float(self.entry2.get())
                    self.ev2 = float(self.ev[1])
                    print(self.ev2)
                    
                    self.ev[2] = float(self.entry3.get()) *.01
                    self.ev3 = float(self.ev[2])
                    print(self.ev3)
                    
                    self.ev[3] = float(self.entry4.get()) *.01
                    self.ev4 = float(self.ev[3])
                    print(self.ev4)
                    
                    self.ev[4] = float(self.entry5.get()) *.01
                    self.ev5 = float(self.ev[4])
                    print(self.ev5)
                    
                    self.ev[5] = float(self.entry6.get()) *.01
                    self.ev6 = float(self.ev[5])
                    print(self.ev6)
                    
                    self.ev[6] = float(self.entry7.get())
                    self.ev7 = float(self.ev[6])
                    print(self.ev7)
                    
                    self.ev[7] = float(self.entry8.get())
                    self.ev8 = float(self.ev[7])
                    print(self.ev8)
                    
                    self.ev[8] = float(self.entry9.get())
                    self.ev9 = float(self.ev[8])
                    print(self.ev9)

                    self.ev[9] = float(self.ev2*0.025*self.ev1)
                    self.ev10 = float(self.ev[9])
                    print(self.ev10)

                    self.pv[0] = float(self.entry1pv.get()) *.01
                    self.pv1 = float(self.pv[0])
                    print(self.pv1)
                    
                    self.pv[1] = float(self.entry2pv.get()) *.01
                    self.pv2 = float(self.pv[1])
                    print(self.pv2)
                    
                    self.pv[2] = float(self.entry3pv.get())  *.01
                    self.pv3 = float(self.pv[2])
                    print(self.pv3)
                    
                    self.pv[3] = float(self.entry4pv.get()) *.01
                    self.pv4 = float(self.pv[3])
                    print(self.pv4)
                    
                    self.pv[4] = float(self.entry5pv.get()) *.01
                    self.pv5 = float(self.pv[4])
                    print(self.pv5)
                    
                    self.pv[5] = float(self.entry6pv.get()) *.01
                    self.pv6 = float(self.pv[5])
                    print(self.pv6)

                elif vector:
                    # Assigning vector values to EV and PV arrays

                    if not self.edge:
                        if self.ev[3]== 15:
                            ev4_cell = 2
                        if self.ev[3]== 18:
                            ev4_cell = 3
                        if self.ev[3]== 21:
                            ev4_cell = 4
                        if self.ev[3]== 24:
                            ev4_cell = 5
                        if self.ev[3]== 27:
                            ev4_cell = 6
                        
                        if self.pv[3] == 8:
                            pv4_cell = 7
                        if self.pv[3] == 12:
                            pv4_cell = 8
                        if self.pv[3] == 16:
                            pv4_cell = 9
                        if self.pv[3] == 20:
                            pv4_cell = 10

                        if self.pv[4] == 77:
                            pv5_cell = 11
                        if self.pv[4] == 83:
                            pv5_cell = 12
                        if self.pv[4] == 89:
                            pv5_cell = 13
                        if self.pv[4] == 85:
                            pv5_cell = 14

                        if self.ev[2] == 40:
                            ev3_cell = 15
                        if self.ev[2] == 55:
                            ev3_cell = 16
                        if self.ev[2] == 70:
                            ev3_cell = 17

                        if self.ev[0] == 60:
                            self.ev1_cell = 18
                        if self.ev[0] == 68:
                            self.ev1_cell = 19
                        if self.ev[0] == 70:
                            self.ev1_cell = 20
                        if self.ev[0] == 85:
                            self.ev1_cell = 21

                        if self.ev[4] == 10:
                            ev5_cell = 22
                        if self.ev[4] == 20:
                            ev5_cell = 23

                        if self.ev[5] == 65:
                            ev6_cell = 24
                        if self.ev[5] == 85:
                            ev6_cell = 25

                        if self.ev[6] == 2:
                            ev7_cell = 26
                        if self.ev[6] == 3:
                            ev7_cell = 27

                    self.ev1 = int(self.ev[0])
                    # print(self.ev1)
                    
                    self.ev2 = float(self.ev[1])
                    # print(self.ev2)
                    
                    self.ev[2] = float(self.ev[2]) *.01
                    self.ev3 = float(self.ev[2])
                    # print(self.ev3)
                    
                    self.ev[3] = float(self.ev[3]) *.01
                    self.ev4 = float(self.ev[3])
                    # print(self.ev4)
                    
                    self.ev[4] = float(self.ev[4]) *.01
                    self.ev5 = float(self.ev[4])
                    # print(self.ev5)
                    
                    self.ev[5] = float(self.ev[5]) *.01
                    self.ev6 = float(self.ev[5])
                    # print(self.ev6)
                    
                    self.ev[6] = float(self.ev[6])
                    self.ev7 = float(self.ev[6])
                    # print(self.ev7)
                    
                    self.ev[7] = float(self.ev[7])
                    self.ev8 = float(self.ev[7])
                    # print(self.ev8)
                    
                    self.ev[8] = float(self.ev[8])
                    self.ev9 = float(self.ev[8])
                    # print(self.ev9)

                    self.ev[9] = float(self.ev2*0.025*self.ev1)
                    self.ev10 = float(self.ev[9])
                    # print(self.ev10)

                    self.pv[0] = float(self.pv[0]) *.01
                    self.pv1 = float(self.pv[0])
                    # print(self.pv1)
                    
                    self.pv[1] = float(self.pv[1]) *.01
                    self.pv2 = float(self.pv[1])
                    # print(self.pv2)
                    
                    self.pv[2] = float(self.pv[2])  *.01
                    self.pv3 = float(self.pv[2])
                    # print(self.pv3)
                    
                    self.pv[3] = float(self.pv[3]) *.01
                    self.pv4 = float(self.pv[3])
                    # print(self.pv4)
                    
                    self.pv[4] = float(self.pv[4]) *.01
                    self.pv5 = float(self.pv[4])
                    # print(self.pv5)
                    
                    self.pv[5] = float(self.pv[5]) *.01
                    self.pv6 = float(self.pv[5])
                    # print(self.pv6)

            
                # print(f'ev1: {self.ev1}')
                for i in range(self.ev1):
                    val1 = sh_user.cell(i+2,1)
                    val1.value = 'user{}'.format(i+1)
                #assigning UsRec9 into the database
                    val2 = sh_user.cell(i+2,10)
                    val2.value = 0
                #assigning UsRec10 into the database
                    val3 = sh_user.cell(i+2,11)
                    val3.value = 0
                #assigning UsRec11 into the database
                    val4 = sh_user.cell(i+2,12)
                    val4.value = self.ev1
                #assigning UsRec12 into the database
                    val5 = sh_user.cell(i+2,13)
                    val5.value = 'yes'
                #assigning UsRec13 into the database
                    val6 = sh_user.cell(i+2,14)
                    val6.value = 0
                wb_user.save(path_user)
                # print('Initial values for UsRec variables set!')

            # PAGE 8,9
            get_period = current_period_list.index(current_period_list[start_iter]) 
            if current_period_list[get_period] == 'Period Data 1':
                for i in range(2):
                    val = sh_system.cell(i+2,3)
                    val.value = self.ev1
                #assigning SyRec2 into the database
                    va2 = sh_system.cell(i+2,4)
                    va2.value = self.ev10/self.ev1
                #assigning SyRec3 into the database
                    va3 = sh_system.cell(i+2,5)
                    va3.value = 0
                #assigning SyRec4 into the database
                    va4 = sh_system.cell(i+2,6)
                    va4.value = 0
                #assigning SyRec5 into the database
                    va5 = sh_system.cell(i+2,7)
                    va5.value = 0
                #assigning SyRec6 into the database
                    va6 = sh_system.cell(i+2,8)
                    va6.value = 0
                #assigning SyRec7 into the database
                    va7 = sh_system.cell(i+2,9)
                    va7.value = 0
                #assigning SyRec8 into the database
                    va8 = sh_system.cell(i+2,10)
                    va8.value = 0
                #assigning SyRec9 into the database
                    val9 = sh_system.cell(i+2,11)
                    val9.value = 0
                #assigning SyRec10 into the database
                    val10 = sh_system.cell(i+2,12)
                    val10.value = 0
                #assigning SyRec11 into the database
                    val11 = sh_system.cell(i+2,13)
                    val11.value = 0
                #assigning SyRec12 into the database
                    val12 = sh_system.cell(i+2,14)
                    val12.value = 0
                #assigning SyRec13 into the database
                    val13 = sh_system.cell(i+2,15)
                    val13.value = 0
                #assigning SyRec14 into the database
                    val14 = sh_system.cell(i+2,16)
                    val14.value = 0
                #assigning SyRec15 into the database
                    val15 = sh_system.cell(i+2,17)
                    val15.value = 0
                #assigning SyRec16 into the database
                    val16 = sh_system.cell(i+2,18)
                    val16.value = 'no'
                #assigning SyRec17 into the database
                    val17 = sh_system.cell(i+2,19)
                    val17.value = 0
                #assigning SyRec18 into the database
                    val18 = sh_system.cell(i+2,20)
                    val18.value = 0
                #assigning SyRec19 into the database
                    val19 = sh_system.cell(i+2,21)
                    val19.value = self.ev10/self.ev1

                for i in range(3,30):
                    val = sh_system.cell(i+2,3)
                    val.value = 0
                #assigning SyRec2 into the database
                    va2 = sh_system.cell(i+2,4)
                    va2.value = 0
                #assigning SyRec3 into the database
                    va3 = sh_system.cell(i+2,5)
                    va3.value = 0
                #assigning SyRec4 into the database
                    va4 = sh_system.cell(i+2,6)
                    va4.value = 0
                #assigning SyRec5 into the database
                    va5 = sh_system.cell(i+2,7)
                    va5.value = 0
                #assigning SyRec6 into the database
                    va6 = sh_system.cell(i+2,8)
                    va6.value = 0
                #assigning SyRec7 into the database
                    va7 = sh_system.cell(i+2,9)
                    va7.value = 0
                #assigning SyRec8 into the database
                    va8 = sh_system.cell(i+2,10)
                    va8.value = 0
                #assigning SyRec9 into the database
                    val9 = sh_system.cell(i+2,11)
                    val9.value = 0
                #assigning SyRec10 into the database
                    val10 = sh_system.cell(i+2,12)
                    val10.value = 0
                #assigning SyRec11 into the database
                    val11 = sh_system.cell(i+2,13)
                    val11.value = 0
                #assigning SyRec12 into the database
                    val12 = sh_system.cell(i+2,14)
                    val12.value = 0
                #assigning SyRec13 into the database
                    val13 = sh_system.cell(i+2,15)
                    val13.value = 0
                #assigning SyRec14 into the database
                    val14 = sh_system.cell(i+2,16)
                    val14.value = 0
                #assigning SyRec15 into the database
                    val15 = sh_system.cell(i+2,17)
                    val15.value = 0
                #assigning SyRec16 into the database
                    val16 = sh_system.cell(i+2,18)
                    val16.value = 0
                #assigning SyRec17 into the database
                    val17 = sh_system.cell(i+2,19)
                    val17.value = 0
                #assigning SyRec18 into the database
                    val18 = sh_system.cell(i+2,20)
                    val18.value = 0
                #assigning SyRec19 into the database
                    val19 = sh_system.cell(i+2,21)
                    val19.value = 0
                wb_system.save(path_system)
                # print('Initial values for SyRec variables set!')
            
                # Subgroup  #FUNCTION FOR SUBGROUP EXECUTION
                self.ev1 = int(self.ev[0])
                step1_ev1 = self.ev1
                step2 = self.ev1/5
                step3 = round(step2/2.3333)
                step4 = step3*5
                step5 = step1_ev1 - step4
                step6 = step5/6
                step7 = round(step6/2)
                step8 = step7*6
                step9 = step5 - step8
                step10 = step9/7
                step11 = int(step10/2)
                step12 = step11*7
                step13 = step9 - step12
                step14 = int(step13/4)
                step15 = step13%4
                if step15 == 0:
                    pass
                if step15 == 1:
                    step3 = step3-1
                    step7 = step7+1
                if step15 == 2:
                    step3 = step3-1
                    step11 = step11+1
                if step15 == 3:
                    step3 = step3-1
                    step14 = step14+2

                #subgroup division code END
                #now assigning number to the group
                #condition checking for group == 4  
                group_num = 1
                group_mem_count = 0
                temp_val_four = step14*4
                four_grp = []
                dep_num = 0
                for i in range(temp_val_four):
                    d = sh_user.cell(i+2,3)
                    d.value = 4
                    a = sh_user.cell(i+2,5)
                    a.value = 4
                    label = sh_user.cell(i+2,2)
                    # label.value = 'D'
                    label.value = group_num
                    four_grp.append(group_num)
                    label_1 = sh_user.cell(i+2,4)
                    # label_1.value = 'D'
                    label_1.value = group_num
                    group_mem_count += 1
                    sh_user.cell(i+2,8).value = 'dependent'
                    dep_num += 1
                    if group_mem_count == 4:
                        group_num += 1
                        group_mem_count = 0
                
                # ('D group assigned!')
                # condition checking for group == 5 
                temp_val_five = step3*5
                for i in range(temp_val_five):
                    d = sh_user.cell(i+temp_val_four+2,3)
                    d.value = 5
                    a = sh_user.cell(i+temp_val_four+2,5)
                    a.value = 5
                    label = sh_user.cell(i+temp_val_four+2,2)
                    # label.value = 'A'
                    label.value = group_num
                    label_1 = sh_user.cell(i+temp_val_four+2,4)
                    # label_1.value = 'A'
                    label_1.value = group_num
                    group_mem_count += 1
                    if group_mem_count == 5:
                        group_num += 1
                        group_mem_count = 0
            
                # condition checking for group == 6
                temp_val_six = step7*6
                for i in range(temp_val_six):
                    d = sh_user.cell(i+temp_val_four+temp_val_five+2,3)
                    d.value = 6
                    a = sh_user.cell(i+temp_val_four+temp_val_five+2,5)
                    a.value = 6
                    label = sh_user.cell(i+temp_val_four+temp_val_five+2,2)
                    # label.value = 'B'
                    label.value = group_num
                    label_1 = sh_user.cell(i+temp_val_four+temp_val_five+2,4)
                    # label_1.value = 'B'
                    label_1.value = group_num
                    group_mem_count += 1
                    if group_mem_count == 6:
                        group_num += 1
                        group_mem_count = 0

                #condition checking for group == 7
                temp_val_seven = step11*7
                for i in range(temp_val_seven):
                    d = sh_user.cell(i+temp_val_four+temp_val_five+temp_val_six+2,3)
                    d.value = 7
                    a = sh_user.cell(i+temp_val_four+temp_val_five+temp_val_six+2,5)
                    a.value = 7
                    label = sh_user.cell(i+temp_val_four+temp_val_five+temp_val_six+2,2)
                    # label.value = 'C'
                    label.value = group_num
                    label_1 = sh_user.cell(i+temp_val_four+temp_val_five+temp_val_six+2,4)
                    # label_1.value = 'C'
                    label_1.value = group_num
                    group_mem_count += 1
                    if group_mem_count == 7:
                        group_num += 1
                        group_mem_count = 0
                wb_user.save(path_user)

                self.UsRec3_dict = {"D": temp_val_four, "A": temp_val_five, "B":  temp_val_six , "C": temp_val_seven}
                checksum = temp_val_four + temp_val_five + temp_val_six + temp_val_seven
                if checksum != self.ev1:
                    raise ValueError(f"Initial group checksum failed: checksum:{checksum} != self.ev1:{self.ev1}")
                # print(self.UsRec3_dict)

                #setting valid to UsRec5
                for i in range(self.ev1):
                    valid_value = sh_user.cell(i+2,6)
                    valid_value.value = 'valid'
                wb_user.save(path_user)
                # print ('group of four members: {}, group of five members: {}, group of six members: {}, group of seven members: {}, Total group: {})'.format(step14, step3, step7, step11, step14*4+step3*5+step7*6+step11*7))


                # RoleAssignment
                self.ev1 = self.ev[0]
                ev4 = self.ev[3]
                ev5 = self.ev[4]
                ev6 = self.ev[5]

                # Assign 'dependent' to equal EV6
                dependent_pct = dep_num/self.ev1
                remaining_pct = ev6 - dependent_pct
                if remaining_pct > 0:
                    unassigned_dep = int(remaining_pct * self.ev1)

                rand_dep_user = sorted(random.sample(range(dep_num+1, self.ev1+1), unassigned_dep))
                
                #ROLE1 
                Role1_list = ['low-morale', 'unity-role']
                #EV 4 = Percentage of honest defectors
                role_ev4 = int(self.ev1*ev4)
                rand_defectors = sorted(random.sample(range(1, self.ev1), role_ev4))
                #EV 5 = Percentage of low-morale members
                role_ev5 = round(self.ev1*ev5)
                low_morale_list = []
                
                if ev5 > 0:
                    while True:
                        n = random.randint(1,self.ev1)
                        if n not in rand_defectors and n not in low_morale_list:
                            low_morale_list.append(n)
                            if len(low_morale_list) == role_ev5 or len(low_morale_list) + len(rand_defectors) == self.ev1:
                            # if len(low_morale_list) == role_ev5:
                                break
                #Remaining members play a unity role
                unity_role = self.ev1 - (role_ev4 - role_ev5)
                #ROLE2
                #percentage of members unwilling to act alone
                role_ev6 = round(self.ev1*ev6)
                #Remaining members play a role of independent
                role_ev12 = self.ev1 - role_ev6
                #assigning UsRec6, ROLE1 values to excel
                assigned_dep = dep_num
                assigned_indep = 0
                for i in range(self.ev1):
                    ur1 = sh_user.cell(i+2,2)
                    UsRec6_init = sh_user.cell(i+2,7)
                    if i+1 in rand_defectors:
                        UsRec6_init.value = 'defector'
                    elif i+1 in low_morale_list:
                        UsRec6_init.value = 'low-morale'
                    
                    UsRec2_init = sh_user.cell(i+2,3)
                    if UsRec2_init.value != 4 and i+1 in rand_dep_user:
                        UsRec7_init = sh_user.cell(i+2,8)
                        UsRec7_init.value =  'dependent'
                        assigned_dep += 1
                    elif UsRec2_init.value != 4 and i+1 not in rand_dep_user:
                        UsRec7_init = sh_user.cell(i+2,8)
                        UsRec7_init.value =  'independent'
                        assigned_indep += 1
                wb_user.save(path_user)
                for i in range(self.ev1):
                    UsRec6_init = sh_user.cell(i+2,7)
                    if UsRec6_init.value != 'defector':
                        if UsRec6_init.value != 'low-morale':
                            UsRec6_init.value = 'unity-role'

                if assigned_dep + assigned_indep != self.ev1:
                    print(f'Dependent/independent assignment error')
                
                wb_user.save(path_user)
                # print('Roles Assigned!')
 
            #################
            ## ___UsFunc1___
            #################
            assign_variables()
            """"
            Pay Stage 1
            USER DEFECTION FUNCTION
            """
                                    
            if current_period_list[start_iter] == 'Period Data 1':
                defector_count = 0
                current_group_num = 1
          
                # Setting defector values in each subgroup
                defected_cache = {}
                defected_subt = {}
                low_morale_cache = []
                lm_def = []
                for i in range(self.ev1):
                    UsRec1 = sh_user.cell(i+2,2)
                    UsRec13 = sh_user.cell(i+2,14)

                    UsRec6 = sh_user.cell(i+2,7)
                    UsRec7 = sh_user.cell(i+2,8)

                    if UsRec6.value == 'defector' or i == self.ev1-1:
                        if current_group_num not in defected_subt:
                            defected_subt[current_group_num] = 0
                        # PATH 1 for dependent
                        if UsRec1.value == current_group_num or i == self.ev1-1:
                            if UsRec6.value == 'defector':
                                defector_count += 1
                                UsRec13.value = defector_count
                                if UsRec7.value == 'dependent':
                                        defected_subt[current_group_num] += 1
                                if UsRec13.value >= self.ev7 or UsRec7.value == 'independent':
                                    # PATH 2 (Part 1 - assigning to cache for UsRec 6 & 13 incrementation in next for loop)
                                    # for independent and dependent
                                    defected_cache[current_group_num] = defector_count
                        if UsRec1.value != current_group_num or i == self.ev1-1:
                            if current_group_num in defected_cache:
                                if defected_cache[current_group_num] < self.ev7:
                                    if defected_subt[current_group_num] != 0:
                                        defected_cache[current_group_num] = defector_count
                                        defected_cache[current_group_num] -= defected_subt[current_group_num]
                                        low_morale_cache.append(current_group_num)
                                        lm_def.append(current_group_num)
                            
                            if defector_count < self.ev7 and current_group_num not in defected_cache and defector_count != 0:
                                # PATH 3 (Part 1 - assigning to cache for UsRec 6 & 7 values in next for loop)
                                low_morale_cache.append(current_group_num)
                            if i != self.ev1-1:
                                defector_count = 1
                                current_group_num = UsRec1.value
                                if current_group_num not in defected_subt:
                                    defected_subt[current_group_num] = 0
                                UsRec13.value = defector_count
                                if UsRec7.value == 'independent':
                                    # PATH 2 (Part 1 - assigning to cache for UsRec 6 & 13 incrementation in next for loop)
                                    # for independent in next group
                                    defected_cache[current_group_num] = defector_count
                                elif UsRec7.value == 'dependent':
                                    defected_subt[current_group_num] += 1
                wb_user.save(path_user)

                # PATH 2 & 3 (Part 2)
                for i in range(self.ev1):
                    UsRec1 = sh_user.cell(i+2,2)
                    UsRec3 = sh_user.cell(i+2,4)
                    UsRec4 = sh_user.cell(i+2,5)
                    UsRec5 = sh_user.cell(i+2,6)
                    UsRec6 = sh_user.cell(i+2,7)
                    UsRec7 = sh_user.cell(i+2,8)
                    UsRec8 = sh_user.cell(i+2,9)
                    UsRec12 = sh_user.cell(i+2,13)
                    UsRec13 = sh_user.cell(i+2,14)

                    if UsRec1.value in defected_cache:
                        # PATH 2 (Part 2)
                        UsRec13.value = defected_cache[UsRec1.value]
                        group_mems = UsRec4.value - UsRec13.value
                        if UsRec6.value == 'defector' and UsRec1.value not in lm_def:
                            # Defectors >= ev7
                            UsRec12.value = 'no'
                            UsRec8.value = 'defected'
                            self.SyRec1_p.value -= 1
                            self.SyRec5_p.value += 1
                            self.SyRec3_p.value += 1
                            
                            for _ in range(self.ev1):
                                ur4 = sh_user.cell(_+2, 5)
                                ur1 = sh_user.cell(_+2, 2)
                                ur3 = sh_user.cell(_+2, 4)
                                ur2 = sh_user.cell(_+2, 3)

                                if ur4.value != 0:
                                    if UsRec3.value == ur3.value:
                                        ur4.value -= 1
                                        # ur4.value = group_mems
                                        if UsRec1.value == ur1.value:
                                            ur2.value -= 1
                            UsRec3.value = 0
                            UsRec4.value = 0
                            UsRec5.value = 'NR'
                            UsRec8.value = 'NR'
                            UsRec12.value = 'NR'
                            wb_user.save(path_user)

                        if UsRec6.value == 'defector' and UsRec1.value in lm_def and UsRec7.value == 'independent':
                            # Defectors < ev7 and independent defectors exist
                            UsRec12.value = 'no'
                            UsRec8.value = 'defected'
                            self.SyRec1_p.value -= 1
                            self.SyRec5_p.value += 1
                            self.SyRec3_p.value += 1
                            
                            for _ in range(self.ev1):
                                ur4 = sh_user.cell(_+2, 5)
                                ur1 = sh_user.cell(_+2, 2)
                                ur3 = sh_user.cell(_+2, 4)
                                ur2 = sh_user.cell(_+2, 3)
                                if ur4.value != 0:
                                    if UsRec3.value == ur3.value:
                                        ur4.value -= 1
                                        # ur4.value = group_mems
                                        if UsRec1.value == ur1.value:
                                            ur2.value -= 1
                            UsRec3.value = 0
                            UsRec4.value = 0
                            UsRec5.value = 'NR'
                            UsRec8.value = 'NR'
                            UsRec12.value = 'NR'
                            wb_user.save(path_user)

                    if UsRec1.value in low_morale_cache and UsRec7.value == 'dependent':
                        # PATH 3 (Part 2)
                        if UsRec1.value in defected_cache:
                            UsRec13.value = defected_cache[UsRec1.value]
                        else:
                            UsRec13.value = 0
                        if UsRec6.value == 'defector':
                            UsRec6.value = 'low-morale'
                wb_user.save(path_user)
                wb_system.save(path_system)  

                _checksum(1, int(counter), 1080)            
                
            #################
            # ___UsFunc2___
            #################
            assign_variables()
            """"
            Pay Stage 2
            USER DEFECTION FUNCTION
            """
            
            if current_period_list[start_iter] != 'Period Data 1':

                slope = (self.pv4 - self.pv2) / (self.pv3 - self.pv1)
                
                SyRec19_prev = sh_system.cell(self.row1-3,21)
                try:
                    a = float(self.SyRec19_p.value)
                    b = float(SyRec19_prev.value)
                    Inc_premium = (a/b) - 1
                except Exception as e:
                    print(e)
                    print(f'SyRec19.value: {self.SyRec19_p.value}')
                    print(f'SyRec19_prev.value: {SyRec19_prev.value}')
                    print(f'row1: {self.row1}')
                    print(f'row1-3: {self.row1-3}')

                valid_users = []
                for i in range(self.ev1):
                    UsRec5 = sh_user.cell(i+2,6)
                    if UsRec5.value == 'valid':
                        valid_users.append(i+2)
                
                if Inc_premium >= self.pv1:
                    # PATH1 
                    skip_percent = (slope*Inc_premium - slope * self.pv1) + self.pv2

                    skip_hash = round(self.SyRec1_p.value * skip_percent)
                    skip_users = random.sample(valid_users, skip_hash)

                    for i in range(self.ev1):
                        index = i+2
                        UsRec12 = sh_user.cell(index,13)
                        if index in skip_users:
                            UsRec12.value = 'no'

                if Inc_premium < self.pv1:
                    try:
                        num = (self.SyRec19_p.value/(float(self.ev10/self.ev1))-1)
                        if num >= self.pv5:
                            skip_hash = round(self.SyRec1_p.value * self.pv6)
                            skip_users = random.sample(valid_users, skip_hash)

                            # rand_skip_users = []
                            # for _ in range(skip_hash):
                            #     n = random.randint(2,self.ev1)
                            #     while True:
                            #         if n in rand_skip_users:
                            #             n = random.randint(2,self.ev1)
                            #         elif n not in rand_skip_users:
                            #             rand_skip_users.append(n)
                            #             break
                            for i in skip_users:
                                UsRec12 = sh_user.cell(i,13)
                                UsRec12.value =  'no'
                            wb_user.save(path_user)
                        if num < self.pv5:
                            # PATH3
                            if self.ev8 == 0:
                                pass
                            if self.ev8 == 1 or self.ev8 == 2 or self.ev8 == 3:
                                self.ev[7] -= 1
                                # valid_users = []
                                # for i in range(self.ev1):
                                #     val = sh_user.cell(i+2,6)
                                #     if val.value == 'valid':
                                #         valid_users.append(i+2)
                                rand_sel = random.choice(valid_users)
                                rand_UsRec12 = sh_user.cell(rand_sel,13)
                                rand_UsRec12.value = 'no'

                    except ZeroDivisionError:
                        pass

                    wb_user.save(path_user)
            
            #################
            # ___SyFunc3___                        #Validate premium function
            #################
            assign_variables()
            """"
            Pay Stage 3
            Validate premium function
            """

            get_period = current_period_list.index(current_period_list[start_iter])
            valid_users = get_valid_users()

            path_1 = []
            path_2 = []
            for i in valid_users:
                
                UsRec1 = sh_user.cell(i, 2)
                UsRec2 = sh_user.cell(i, 3)
                UsRec3 = sh_user.cell(i, 4)
                UsRec4 = sh_user.cell(i, 5)
                UsRec5 = sh_user.cell(i, 6)
                UsRec8 = sh_user.cell(i, 9)
                UsRec12 = sh_user.cell(i, 13)
                
                if UsRec12.value == 'no':
                    UsRec8.value = 'skipped'
                    path_1.append(i)
                    self.SyRec1_p.value -= 1
                    self.SyRec5_p.value += 1 # potential incorrect copying or adding
                    
                    for _ in range(self.ev1):
                        ur4 = sh_user.cell(_+2, 5)
                        ur1 = sh_user.cell(_+2, 2)
                        ur3 = sh_user.cell(_+2, 4)
                        ur2 = sh_user.cell(_+2, 3)
                        
                        if ur4.value != 0:
                            if UsRec3.value == ur3.value:
                                ur4.value -= 1
                                if UsRec1.value == ur1.value:
                                    ur2.value -= 1

                    UsRec8.value = "NR"
                    UsRec3.value = 0
                    UsRec4.value = 0
                    UsRec5.value = "NR"
                    UsRec12.value = "NR"
                    wb_user.save(path_user)

                elif UsRec12.value == 'yes':
                    UsRec8.value = 'paid'
                    
            wb_user.save(path_user)
            wb_system.save(path_system)

            _checksum(3, int(counter), 1265)
            _checksum_sr1(self.SyRec1_p.value, 3, int(counter), 1265)
            
            #################
            # ___SyFunc4___
            #################
            assign_variables()
            """"
            Pay Stage 4
            Invalidate subgroup function
            """

            get_period = current_period_list.index(current_period_list[start_iter])
            _path = 0

            for i in range(self.ev1):    
                ur3 = sh_user.cell(i+2, 4)
                ur4 = sh_user.cell(i+2, 5)
                ur8 = sh_user.cell(i+2, 9)
                ur5 = sh_user.cell(i+2, 6)
                ur10 = sh_user.cell(i+2, 11)
                ur11 = sh_user.cell(i+2, 12)
                
                
                if ur4.value == 1 or ur4.value == 2 or ur4.value == 3:
                    if ur8.value == 'paid':
                        #UsRec8 = 'paid-invalid'
                        ur8.value = 'paid-invalid'
                        # UsRec5 = 'invalid'
                        ur5.value = 'invalid'
                        ur10.value = ur11.value
                        
                        self.SyRec6_p.value += 1 
            wb_system.save(path_system)
            wb_user.save(path_user)
            
            assign_variables()
            if current_period_list[get_period] == 'Period Data 1':
                self.SyRec1_f.value = self.SyRec1_p.value
                self.SyRec1_r.value = self.SyRec1_p.value
                
                # SyRec2 = sh_system.cell(2,4)
                # SyRec2_f = sh_system.cell(3,4)
                # SyRec2_f.value = SyRec2.value
                # SyRec2_r = sh_system.cell(4,4)
                # SyRec2_r.value = SyRec2.value
                self.SyRec2_f.value = self.SyRec2_p.value
                self.SyRec2_r.value = self.SyRec2_p.value
                
                self.SyRec3_f.value = self.SyRec3_p.value
                self.SyRec3_r.value = self.SyRec3_p.value
                
                self.SyRec4_f.value = self.SyRec4_p.value
                self.SyRec4_r.value = self.SyRec4_p.value
                
                self.SyRec5_f.value = self.SyRec5_p.value
                self.SyRec5_r.value = self.SyRec5_p.value
                
                self.SyRec6_f.value = self.SyRec6_p.value
                self.SyRec6_r.value = self.SyRec6_p.value
                
                self.SyRec7_f.value = self.SyRec7_p.value
                self.SyRec7_r.value = self.SyRec7_p.value
                
                self.SyRec8_f.value = self.SyRec8_p.value
                self.SyRec8_r.value = self.SyRec8_p.value
                
                self.SyRec9_f.value = self.SyRec9_p.value
                self.SyRec9_r.value = self.SyRec9_p.value
                
                self.SyRec10_f.value= self.SyRec10_p.value
                self.SyRec10_r.value = self.SyRec10_p.value
                
                self.SyRec11_f.value= self.SyRec11_p.value
                self.SyRec11_r.value = self.SyRec11_p.value
                
                self.SyRec12_f.value= self.SyRec12_p.value
                self.SyRec12_r.value = self.SyRec12_p.value
                
                self.SyRec13_f.value= self.SyRec13_p.value
                self.SyRec13_r.value = self.SyRec13_p.value
                
                self.SyRec14_f.value= self.SyRec14_p.value
                self.SyRec14_r.value = self.SyRec14_p.value
                
                self.SyRec15_f.value= self.SyRec15_p.value
                self.SyRec15_r.value = self.SyRec15_p.value
                
                self.SyRec16_f.value= self.SyRec16_p.value
                self.SyRec16_r.value = self.SyRec16_p.value
                
                self.SyRec17_f.value= self.SyRec17_p.value
                self.SyRec17_r.value = self.SyRec17_p.value
                
                self.SyRec18_f.value= self.SyRec18_p.value
                self.SyRec18_r.value = self.SyRec18_p.value
            
                self.SyRec19_f.value= self.SyRec19_p.value
                self.SyRec19_r.value = self.SyRec19_p.value

                #################
                # ___SyFunc5___
                #################
                assign_variables()
                self.SyRec9_f.value = self.SyRec3_f.value * self.SyRec19_f.value
                sh_system.cell(4,11).value = self.SyRec9_f.value
                wb_system.save(path_system)
            _checksum(4, int(counter), 1480)
            assign_variables()
            if current_period_list[get_period] != 'Period Data 1':

                self.SyRec1_r.value= self.SyRec1_p.value
                # elf.SyRec1 = sh_system.cell(row1,3)
                # self.SyRec1_r = sh_system.cell(row3,3)
                # self.SyRec1_r.value= self.SyRec1.value
                
                self.SyRec2_r.value= self.SyRec2_p.value
                
                self.SyRec3_r.value= self.SyRec3_p.value
                
                self.SyRec4_r.value= self.SyRec4_p.value
                
                self.SyRec5_r.value= self.SyRec5_p.value
                
                self.SyRec6_r.value= self.SyRec6_p.value
                
                self.SyRec7_r.value= self.SyRec7_p.value
                
                self.SyRec8_r.value= self.SyRec8_p.value
                
                self.SyRec9_r.value= self.SyRec9_p.value
                
                self.SyRec10_r.value= self.SyRec10_p.value
                
                self.SyRec11_r.value= self.SyRec11_p.value
                
                self.SyRec12_r.value= self.SyRec12_p.value
                
                self.SyRec13_r.value= self.SyRec13_p.value
                
                self.SyRec14_r.value= self.SyRec14_p.value
                
                self.SyRec15_r.value= self.SyRec15_p.value
                
                self.SyRec16_r.value= self.SyRec16_p.value
                
                self.SyRec17_r.value= self.SyRec17_p.value
                
                self.SyRec18_r.value= self.SyRec18_p.value
            
                self.SyRec19_r.value= self.SyRec19_p.value

            #################                  
            # __SyFunc6__            #User quit function
            #################
            assign_variables()
            """"
            Reorg Stage 1
            """

            low_morale_users = get_select_users('low-morale', 6)

            for i in range(self.ev1):
                _path = 0
                UsRec1 = sh_user.cell(i+2, 2)
                UsRec2 = sh_user.cell(i+2, 3)
                UsRec3 = sh_user.cell(i+2, 4)
                UsRec4 = sh_user.cell(i+2, 5)
                UsRec5 = sh_user.cell(i+2, 6)
                UsRec6 = sh_user.cell(i+2, 7)
                UsRec7 = sh_user.cell(i+2, 8)
                UsRec8 = sh_user.cell(i+2, 9)
                UsRec9 = sh_user.cell(i+2, 10)
                UsRec12 = sh_user.cell(i+2, 13)
                # 1 2 3 2 4
                if UsRec8.value == 'paid-invalid':
                    if UsRec6.value == 'low-morale':
                        # Path 1
                        prob = random.uniform(0,1)
                        if prob >= self.ev9:
                            _path = 3
                        elif prob < self.ev9:
                            _path = 2

                        if _path == 3:
                            if UsRec7.value == 'independent' or UsRec2.value >= 2:
                                #path4 
                                self.SyRec8_r.value += 1
                            else:
                                _path = 2

                            # wb_system.save(path_system)  
                        if _path == 2:
                            UsRec8.value = 'quit'
                            self.SyRec1_r.value -= 1
                            self.SyRec7_r.value += 1
                            for _i in range(self.ev1):
                                ur4 = sh_user.cell(_i+2, 5)
                                ur3 = sh_user.cell(_i+2, 4)
                                ur2 = sh_user.cell(_i+2, 3)
                                ur1 = sh_user.cell(_i+2, 2)

                                if ur4.value != 0:
                                    if ur3.value == UsRec3.value:
                                        ur4.value -= 1
                                        if ur1.value == UsRec1.value:
                                            ur2.value -= 1
                            UsRec8.value = "NR"
                            UsRec3.value = 0
                            UsRec4.value = 0
                            UsRec5.value = "NR"
                            UsRec12.value = "NR"
            wb_system.save(path_system)
            wb_user.save(path_user)

            for i in range(self.ev1):
                _path = 0
                UsRec1 = sh_user.cell(i+2, 2)
                UsRec2 = sh_user.cell(i+2, 3)
                UsRec3 = sh_user.cell(i+2, 4)
                UsRec4 = sh_user.cell(i+2, 5)
                UsRec5 = sh_user.cell(i+2, 6)
                UsRec6 = sh_user.cell(i+2, 7)
                UsRec7 = sh_user.cell(i+2, 8)
                UsRec8 = sh_user.cell(i+2, 9)
                UsRec9 = sh_user.cell(i+2, 10)
                UsRec12 = sh_user.cell(i+2, 13)
                # 1 2 3 2 4
                if UsRec8.value == 'paid-invalid':
                    if UsRec6.value != 'low-morale':
                        if UsRec7.value == 'independent' or UsRec2.value >= 2:
                            #path4 
                            self.SyRec8_r.value += 1
                        else:
                            _path = 2

                        # wb_system.save(path_system)  
                        if _path == 2:
                            UsRec8.value = 'quit'
                            self.SyRec1_r.value -= 1
                            self.SyRec7_r.value += 1
                            for _i in range(self.ev1):
                                ur4 = sh_user.cell(_i+2, 5)
                                ur3 = sh_user.cell(_i+2, 4)
                                ur2 = sh_user.cell(_i+2, 3)
                                ur1 = sh_user.cell(_i+2, 2)

                                if ur4.value != 0:
                                    if ur3.value == UsRec3.value:
                                        ur4.value -= 1
                                        if ur1.value == UsRec1.value:
                                            ur2.value -= 1
                            UsRec8.value = "NR"
                            UsRec3.value = 0
                            UsRec4.value = 0
                            UsRec5.value = "NR"
                            UsRec12.value = "NR"
            wb_system.save(path_system)
            wb_user.save(path_user)

            _checksum(6, int(counter), 1520)
            _checksum_sr1(self.SyRec1_r.value, 6, int(counter), 1520)

            #################        
            # ___SyFunc7___
            #################
            assign_variables()
            """"
            Reorg Stage 2
            """

            _path = 0
            loop_reset = False
            invalid_loop = 0  # UsRec4 for group absorbing invalid member not reassigned twice
            found_subgrp = 0
            pass_over = ["defected","skipped","quit","NR"]
            for i in range(self.ev1):
                ur8 = sh_user.cell(i+2, 9) 
                ur4 = sh_user.cell(i+2, 5)
                if ur8.value =='paid-invalid':

                    if ur4.value == 1:
                        ur1 = sh_user.cell(i+2, 2)
                        base_ur4 = ur4.value
                        invalid_loop += 1

                        # Chnage ur4 values of group absorbing invalid member
                        if invalid_loop == 1:
                            old_ur4 = 0
                            new_ur4 = 0 # Used to set subgroup so not override values in edge cases
                            for _i in range(self.ev1):
                                ur4_sub = sh_user.cell(_i+2, 5)
                                ur3_sub = sh_user.cell(_i+2, 4)
                                ur5_sub = sh_user.cell(_i+2, 6)
                                ur8_sub = sh_user.cell(_i+2, 9)
                                if new_ur4 == 0:
                                    if ur4_sub.value == 6 and ur8_sub.value not in pass_over and ur5_sub.value == 'valid':
                                        found_subgrp = ur3_sub.value
                                        ur4_sub.value = 7
                                        old_ur4 = 6
                                        new_ur4 = 7
                                        _path = 1
                                elif ur4_sub.value == old_ur4  and ur8_sub.value not in pass_over and ur5_sub.value == 'valid':
                                    if ur3_sub.value == found_subgrp:
                                        ur4_sub.value = new_ur4

                            if _path != 1:
                                for _i in range(self.ev1):
                                    ur4_sub = sh_user.cell(_i+2, 5)
                                    ur3_sub = sh_user.cell(_i+2, 4)
                                    ur5_sub = sh_user.cell(_i+2, 6)
                                    ur8_sub = sh_user.cell(_i+2, 9)
                                    if new_ur4 == 0:
                                        if ur4_sub.value == 5 and ur8_sub.value not in pass_over and ur5_sub.value == 'valid':
                                                found_subgrp = ur3_sub.value
                                                ur4_sub.value = 6
                                                old_ur4 = 5
                                                new_ur4 = 6
                                    elif ur4_sub.value == old_ur4 and ur8_sub.value not in pass_over and ur5_sub.value == 'valid':
                                        if ur3_sub.value == found_subgrp:
                                            ur4_sub.value = new_ur4

                        if invalid_loop == base_ur4:
                            loop_reset = True
                        ur4.value = new_ur4
                        ur3 = sh_user.cell(i+2, 4)
                        ur3.value = found_subgrp #!!! not referenced
                        ur5= sh_user.cell(i+2, 6)
                        ur5.value = 'valid'
                        ur8.value = 'reorg'
                        ur9= sh_user.cell(i+2, 10)
                        ur9.value = ur9.value + 1

                        if loop_reset:
                            invalid_loop = 0
                            loop_reset = False
            wb_user.save(path_user)
            _checksum(7, int(counter), 1761)
            
            loop_reset = False
            invalid_loop = 0  # UsRec4 for group absorbing invalid member not reassigned twice
            _path = 0
            found_subgrp = 0
            reorg_cache = {}
            for i in range(self.ev1):
                ur8 = sh_user.cell(i+2, 9)
                ur3 = sh_user.cell(i+2, 4)
                ur4 = sh_user.cell(i+2, 5)
                if ur8.value =='paid-invalid':
                    ur1 = sh_user.cell(i+2, 2)

                    if ur4.value == 2:
                        base_ur4 = ur4.value
                        invalid_loop += 1

                        # Change ur4 values of group absorbing invalid member
                        if invalid_loop == 1 and ur3.value not in reorg_cache:
                            old_ur4 = 0
                            new_ur4 = 0 # Used to set subgroup so not override values in edge cases
                            for _i in range(self.ev1):
                                ur4_sub = sh_user.cell(_i+2, 5)
                                ur3_sub = sh_user.cell(_i+2, 4)
                                ur5_sub = sh_user.cell(_i+2, 6)
                                ur8_sub = sh_user.cell(_i+2, 9)
                                if new_ur4 == 0:
                                    if ur4_sub.value == 5 and ur8_sub.value not in pass_over:# and ur5_sub.value == 'valid':
                                        found_subgrp = ur3_sub.value
                                        ur4_sub.value = 7
                                        old_ur4 = 5
                                        new_ur4 = 7
                                    elif ur4_sub.value == 4 and ur8_sub.value not in pass_over:# and ur5_sub.value == 'valid':
                                        found_subgrp = ur3_sub.value
                                        ur4_sub.value = 6
                                        old_ur4 = 4
                                        new_ur4 = 6
                                elif ur4_sub.value == old_ur4 and ur8_sub.value not in pass_over:# and ur5_sub.value == 'valid':
                                    if ur3_sub.value == found_subgrp:
                                        ur4_sub.value = new_ur4

                            reorg_cache.update({ur3.value:found_subgrp})
                        
                        if invalid_loop == base_ur4:
                            loop_reset = True
                        if ur3.value in reorg_cache:
                            found_subgrp = reorg_cache[ur3.value]
                            loop_reset = True
                        ur4.value = new_ur4
                        ur3.value = found_subgrp
                        ur5= sh_user.cell(i+2, 6)
                        ur5.value = 'valid'
                        ur8.value = 'reorg'
                        ur9= sh_user.cell(i+2, 10)
                        ur9.value = ur9.value + 1
                        
                        if loop_reset:
                            invalid_loop = 0
                            loop_reset = False
            wb_user.save(path_user)
            _checksum(7, int(counter), 1761)                
            
            loop_reset = False
            invalid_loop = 0  # UsRec4 for group absorbing invalid member not reassigned twice
            _path = 0
            found_subgrp = 0
            reorg_cache = {}
            for i in range(self.ev1):
                ur8 = sh_user.cell(i+2, 9)
                ur4 = sh_user.cell(i+2, 5)
                
                if ur8.value =='paid-invalid':
                    ur1 = sh_user.cell(i+2, 2)
                    ur3 = sh_user.cell(i+2, 4)

                    if ur4.value == 3:
                        base_ur4 = ur4.value
                        invalid_loop += 1

                        # Change ur4 values of group absorbing invalid member
                        if invalid_loop == 1 and ur3.value not in reorg_cache:
                            grp_found = 0
                            old_ur4 = 0
                            new_ur4 = 0 # Used to set subgroup so not override values in edge cases
                            for _i in range(self.ev1):
                                
                                ur4_sub = sh_user.cell(_i+2, 5)
                                ur3_sub = sh_user.cell(_i+2, 4)
                                ur5_sub = sh_user.cell(_i+2, 6)
                                ur8_sub = sh_user.cell(_i+2, 9)
                                if new_ur4 == 0:
                                    if ur4_sub.value == 3 and ur3_sub.value != ur3.value and ur8_sub.value not in pass_over:# and ur5_sub.value == 'valid':
                                        found_subgrp = ur3_sub.value
                                        ur4_sub.value = 6
                                        old_ur4 = 3
                                        new_ur4 = 6
                                        _path = 2
                                        ur5_sub.value = 'valid'
                                        ur8_sub.value = 'reorg'
                                        ur9_sub = sh_user.cell(_i+2, 10)
                                        ur9_sub.value = ur9_sub.value + 1
                                        grp_found = ur3_sub.value
                                elif ur4_sub.value == old_ur4 and ur8_sub.value not in pass_over: #and ur5_sub.value == 'valid':
                                    if ur3_sub.value == found_subgrp:
                                        ur4_sub.value = new_ur4
                                        ur5_sub.value = 'valid'
                                        ur8_sub.value = 'reorg'
                                        ur9_sub = sh_user.cell(_i+2, 10)
                                        ur9_sub.value = ur9_sub.value + 1

                            if grp_found != 0:
                                reorg_cache.update({ur3.value:found_subgrp})

                            if grp_found == 0:
                                for _i in range(self.ev1):
                                    
                                    ur4_sub = sh_user.cell(_i+2, 5)
                                    ur3_sub = sh_user.cell(_i+2, 4)
                                    ur5_sub = sh_user.cell(_i+2, 6)
                                    ur8_sub = sh_user.cell(_i+2, 9)
                                    if new_ur4 == 0:
                                        if ur4_sub.value == 4 and ur8_sub.value not in pass_over: #and ur5_sub.value == 'valid':
                                            found_subgrp = ur3_sub.value
                                            old_subgroup = ur3.value
                                            ur4_sub.value = 7
                                            old_ur4 = 4
                                            new_ur4 = 7
                                            _path = 1
                                            
                                    elif ur4_sub.value == old_ur4 and ur8_sub.value not in pass_over:# and ur5_sub.value == 'valid':
                                        if ur3_sub.value == found_subgrp:
                                            ur4_sub.value = new_ur4
                                reorg_cache.update({ur3.value:found_subgrp})

                        if invalid_loop == base_ur4:
                            loop_reset = True

                        if ur3.value in reorg_cache:
                            found_subgrp = reorg_cache[ur3.value]
                            loop_reset = True

                        if _path == 2:
                            ur4.value = new_ur4
                            ur3.value = found_subgrp
                            ur5= sh_user.cell(i+2, 6)
                            ur5.value = 'valid'
                            ur8.value = 'reorg'
                            ur9= sh_user.cell(i+2, 10)
                            ur9.value = ur9.value + 1

                        if _path == 1:
                            ur4.value = new_ur4
                            ur3 = sh_user.cell(i+2, 4)
                            ur3.value = found_subgrp
                            ur5= sh_user.cell(i+2, 6)
                            ur5.value = 'valid'
                            ur8.value = 'reorg'
                            ur9= sh_user.cell(i+2, 10)
                            ur9.value = ur9.value + 1
                        
                        if loop_reset:
                            invalid_loop = 0
                            loop_reset = False

            wb_user.save(path_user)

            _checksum(7, int(counter), 1761)

            #################
            # ___SyFunc8___    Claims / refunds function
            #################
            assign_variables()
            """"
            Reorg Stage 4
            """
            
            prob = round(random.uniform(0,1),2)
            if self.ev3 > prob:
                self.SyRec16_r.value = 'yes'
            elif self.ev3 < prob:
                self.SyRec16_r.value = "no"
                self.SyRec17_r.value = self.SyRec2_r.value
            wb_system.save(path_system)
            
            #################
            # ___SyFunc8.5___
            #################
            assign_variables()
            """"
            Reorg Stage 4.5
            """
            self.SyRec11_r.value = self.SyRec5_r.value * self.SyRec19_r.value
            self.SyRec13_r.value = self.SyRec6_r.value * self.SyRec19_r.value
            wb_system.save(path_system)                

            #################     
            # ___SyFunc9___
            #################
            assign_variables()
            """"
            Reorg Stage 5
            """
            
            try:
                self.SyRec2_r.value = float(self.ev10)/self.SyRec1_r.value
            except ZeroDivisionError:
                pass

            try:
                self.SyRec14_r.value = self.SyRec9_r.value + self.SyRec11_r.value + self.SyRec13_r.value
            except ZeroDivisionError:
                pass

            try:
                self.SyRec15_r.value = self.SyRec14_r.value / self.SyRec1_r.value
            except ZeroDivisionError:
                pass
            
            for i in range(self.ev1):
                us10 = sh_user.cell(i+2, 11)
                us11 = sh_user.cell(i+2, 12)
                if us10.value != 0:
                    us11.value = self.SyRec2_r.value + self.SyRec15_r.value - us10.value
                    us10.value = 0
                else:
                    us11.value = self.SyRec2_r.value + self.SyRec15_r.value - self.SyRec18_r.value
            self.SyRec19_r.value = self.SyRec2_r.value + self.SyRec15_r.value
            wb_system.save(path_system)
            wb_user.save(path_user)            

            #################
            # ___SyFunc11___
            #################
            assign_variables()
            """"
            Reorg Stage 7
            """
            _path = 0
            get_period = current_period_list.index(current_period_list[start_iter])        

            if current_period_list[get_period] != 'Period Data 10':
                total = self.SyRec3_r.value+self.SyRec5_r.value+self.SyRec7_r.value

                if total > 0:
                    _path = 1
                elif total == 0:
                    _path = 2

                if _path == 1:
            
                    if get_period==0:
                        pay_row=2
                        reorg_row=4
                        new_pay_row = 5
                    if get_period==1:
                        pay_row=5
                        reorg_row=7
                        new_pay_row = 8
                    if get_period==2:
                        pay_row=8
                        reorg_row=10
                        new_pay_row = 11
                    if get_period==3:
                        pay_row=11
                        reorg_row=13
                        new_pay_row = 14
                    if get_period==4:
                        pay_row=14
                        reorg_row=16
                        new_pay_row = 17
                    if get_period==5:
                        pay_row=17
                        reorg_row=19
                        new_pay_row = 20
                    if get_period==6:
                        pay_row=20
                        reorg_row=22
                        new_pay_row = 23
                    if get_period==7:
                        pay_row=23
                        reorg_row=25
                        new_pay_row = 26
                    if get_period==8:
                        pay_row=26
                        reorg_row=28
                        new_pay_row = 29
                    if get_period==9:
                        pay_row=29
                        reorg_row=31

                    #copying values of previous to current
                    SyRec1_r = sh_system.cell(reorg_row,3)
                    SyRec1_new_p = sh_system.cell(new_pay_row,3)
                    # SyRec1_r.value = SyRec1_p.value
                    SyRec1_new_p.value = SyRec1_r.value

                    SyRec2_r = sh_system.cell(reorg_row,4)
                    SyRec2_new_p = sh_system.cell(new_pay_row,4)
                    # SyRec2_r.value = SyRec2_p.value
                    SyRec2_new_p.value = SyRec2_r.value

                    SyRec3_r = sh_system.cell(reorg_row,5)
                    SyRec3_new_p = sh_system.cell(new_pay_row,5)
                    # SyRec3_r.value = SyRec3_p.value
                    SyRec3_new_p.value = SyRec3_r.value
                    
                    SyRec4_r = sh_system.cell(reorg_row,6)
                    SyRec4_new_p = sh_system.cell(new_pay_row,6)
                    # SyRec4_r.value = SyRec4_p.value
                    SyRec4_new_p.value = SyRec4_r.value
                    
                    SyRec5_r = sh_system.cell(reorg_row,7)
                    SyRec5_new_p = sh_system.cell(new_pay_row,7)
                    # SyRec5_r.value = SyRec5_p.value
                    SyRec5_new_p.value = SyRec5_r.value

                    SyRec6_r = sh_system.cell(reorg_row,8)
                    SyRec6_new_p = sh_system.cell(new_pay_row,8)
                    # SyRec6_r.value = SyRec6_p.value
                    SyRec6_new_p.value = SyRec6_r.value
                    
                    SyRec7_r = sh_system.cell(reorg_row,9)
                    SyRec7_new_p = sh_system.cell(new_pay_row,9)
                    # SyRec7_r.value = SyRec7_p.value
                    SyRec7_new_p.value = SyRec7_r.value
                    
                    SyRec8_r = sh_system.cell(reorg_row,10)
                    SyRec8_new_p = sh_system.cell(new_pay_row,10)
                    # SyRec8_r.value = SyRec8_p.value
                    SyRec8_new_p.value = SyRec8_r.value
                    
                    SyRec9_r = sh_system.cell(reorg_row,11)
                    SyRec9_new_p = sh_system.cell(new_pay_row,11)
                    # SyRec9_r.value = SyRec9_p.value
                    SyRec9_new_p.value = SyRec9_r.value
                    
                    SyRec10_r = sh_system.cell(reorg_row,12)
                    SyRec10_new_p = sh_system.cell(new_pay_row,12)
                    # SyRec10_r.value = SyRec10_p.value
                    SyRec10_new_p.value = SyRec10_r.value
                    
                    SyRec11_r = sh_system.cell(reorg_row,13)
                    SyRec11_new_p = sh_system.cell(new_pay_row,13)
                    # SyRec11_r.value = SyRec11_p.value
                    SyRec11_new_p.value = SyRec11_r.value
                    
                    SyRec12_r = sh_system.cell(reorg_row,14)
                    SyRec12_new_p = sh_system.cell(new_pay_row,14)
                    # SyRec12_r.value = SyRec12_p.value
                    SyRec12_new_p.value = SyRec12_r.value
                    
                    SyRec13_r = sh_system.cell(reorg_row,15)
                    SyRec13_new_p = sh_system.cell(new_pay_row,15)
                    # SyRec13_r.value = SyRec13_p.value
                    SyRec13_new_p.value = SyRec13_r.value
                    
                    SyRec14_r = sh_system.cell(reorg_row,16)
                    SyRec14_new_p = sh_system.cell(new_pay_row,16)
                    # SyRec14_r.value = SyRec14_p.value
                    SyRec14_new_p.value = SyRec14_r.value
                    
                    SyRec15_r = sh_system.cell(reorg_row,17)
                    SyRec15_new_p = sh_system.cell(new_pay_row,17)
                    # SyRec15_r.value = SyRec15_p.value
                    SyRec15_new_p.value = SyRec15_r.value
                    
                    SyRec16_r = sh_system.cell(reorg_row,18)
                    SyRec16_new_p = sh_system.cell(new_pay_row,18)
                    # SyRec16_r.value = SyRec16_p.value
                    SyRec16_new_p.value = SyRec16_r.value
                    
                    SyRec17_r = sh_system.cell(reorg_row,19)
                    SyRec17_new_p = sh_system.cell(new_pay_row,19)
                    # SyRec17_r.value = SyRec17_p.value
                    SyRec17_new_p.value = SyRec17_r.value
                    
                    SyRec18_r = sh_system.cell(reorg_row,20)
                    SyRec18_new_p = sh_system.cell(new_pay_row,20)
                    # SyRec18_r.value = SyRec18_p.value
                    SyRec18_new_p.value = SyRec18_r.value
                    
                    SyRec19_r = sh_system.cell(reorg_row,21)
                    SyRec19_new_p = sh_system.cell(new_pay_row,21)
                    # SyRec19_r.value = SyRec19_p.value
                    SyRec19_new_p.value = SyRec19_r.value
                    
                    wb_system.save(path_system)
                    # Overwriting values in new row

                    SyRec18_new_p.value = SyRec17_new_p.value
                    wb_system.save(path_system)
                    SyRec11_new_p.value = 0
                    SyRec13_new_p.value = 0
                    SyRec14_new_p.value = 0
                    SyRec15_new_p.value = 0
                    SyRec9_new_p.value = 0
                    SyRec10_new_p.value = 0
                    SyRec17_new_p.value = 0 
                    SyRec3_new_p.value = 0
                    SyRec5_new_p.value = 0 
                    SyRec6_new_p.value = 0
                    wb_system.save(path_system)

                    _checksum_sr1(SyRec1_new_p.value, 11, int(counter), 2089)
                    _checksum(11,int(counter),2088)
                    
            
            #logging to log file  
            if current_period_list[get_period] == 'Period Data 10' or _path == 2:
                # print(f'Run complete, logging simulation results')
                try:
                    if not self._matrix:
                        x = datetime.now()
                        file1_path = os.path.join(root_sys, 'result.txt')
                        
                        file1 = open(file1_path, 'w')
                        # file1 = open("C:\\Tandapay\\LOGS\\myfile.txt", 'w')
                        log1 = f'{self.ev1} is the number of members at the start of the simulation\n'
                        log2 = f'{self.SyRec1_r.value} is the number of valid members remaining at the end of the simulation\n'
                        log3 = f'{round(((self.ev1 - self.SyRec1_r.value) / self.ev1)*100,2)}% of policyholders left the group by end of simulation\n'
                        log4 = f'{round(sh_system.cell(2,21).value)} was the initial premium members were asked to pay.\n'
                        log5 = f'{self.SyRec19_f.value} is the final premium members were asked to pay.\n'
                        log6 = f'Premiums increased by {round((self.SyRec19_f.value / sh_system.cell(2,21).value)*100,2)}% by end of simulation\n'
                        log7 = f'self.SyRec 3 (period 0 finalize) = {sh_system.cell(3,5).value}\n'
                        log8 = f'{self.ev4*100}% of policyholders who were assigned to defect\n'
                        log9 = (sh_system.cell(3,5).value/self.ev1)*100
                        log9 = f'{round(log9,2)}% of policyholders who actually defected\n'
                        log10 = f'{(self.pv5)*100}% was the initial collapse threshold set for PV 5\n'
                    
                        L=[log1, log2, log3, log4, log5, log6, log7, log8, log9, log10]
                        for _log in L:
                            print(_log)
                        file1.writelines(L)
                        file1.close()

                        #copying excel file to another location
                    
                        original_sys = os.path.join(root_sys, '1 System Database.xlsx')
                        original_us = os.path.join(root_sys, '2 User Database.xlsx')

                        original_sys = os.path.join(root_sys, '1 System Database.xlsx')
                        original_us = os.path.join(root_sys, '2 User Database.xlsx')
                        edge_sys = os.path.join(root_sys, '1 System Database.xlsx')
                        edge_us = os.path.join(root_sys, '2 User Database.xlsx')

                        target = os.path.join(root_sys, 'LOGS', '1 System Database temp copy.xlsx')
                        print(original_sys + ' -> ' + target)
                        # original = r"C:\Tandapay\1 System Database.xlsx"
                        # target = r"C:\Tandapay\LOGS\1 System Database temp copy.xlsx"
                        shutil.copyfile(original_sys, target)

                        if self.edge:
                            shutil.copyfile(original_sys, edge_sys)
                            shutil.copyfile(original_us, edge_us)

                    elif self._matrix:
                        log1 = self.ev1
                        log2 = self.SyRec1_r.value
                        log3 = round(((self.ev1 - self.SyRec1_p.value) / self.ev1)*100,2)
                        log4 = round(sh_system.cell(2,21).value)
                        log5 = self.SyRec19_f.value
                        log6 = round((self.SyRec19_f.value / sh_system.cell(2,21).value)*100,2)
                        log7 = sh_system.cell(3,5).value
                        log8 = self.ev4*100
                        log9 = (sh_system.cell(3,5).value/self.ev1)*100
                        log9 = round(log9,2)
                        log10 = (self.pv5)*100
                    
                        L=[log1, log2, log3, log4, log5, log6, log7, log8, log9, log10]

                        # self.SyRec1_p = sh_system.cell(self.row3,3)
                        
                        if self.SyRec1_f.value < self.ev1/2:
                            collapse = 1
                        else:
                            collapse = 0
                        
                        row = self.run + 1

                        if not self.edge:
                            matrix_var_sh.cell(row, 1).value = self.run
                            matrix_sys_log.cell(row, 1).value = self.run

                            matrix_var_sh_cols = [ev4_cell,pv4_cell,pv5_cell,ev3_cell,self.ev1_cell,ev5_cell,ev6_cell,ev7_cell]

                            for col in matrix_var_sh_cols:
                                matrix_var_sh.cell(row, col).value = collapse

                            for index, var in enumerate(self.start_ev):
                                log_index = index+2
                                matrix_sys_log.cell(row, log_index).value = var
                                if log_index == 8:
                                    break
                            for index, var in enumerate(self.start_pv):
                                log2_index = log_index + index + 1
                                matrix_sys_log.cell(row, log2_index).value = var

                            for index, var in enumerate(L):
                                log3_index = log2_index + index + 1
                                matrix_sys_log.cell(row, log3_index).value = var

                            matrix_wb.save(path_matrix)
                            return
                        elif self.edge:
                            file1_path = os.path.join(root_sys, 'LOGS', str(self.run) +'_'+'log.txt')
                        
                            file1 = open(file1_path, 'w')
                            # file1 = open("C:\\Tandapay\\LOGS\\myfile.txt", 'w')
                            log1 = f'{self.ev1} is the number of members at the start of the simulation\n'
                            log2 = f'{self.SyRec1_r.value} is the number of valid members remaining at the end of the simulation\n'
                            log3 = f'{round(((self.ev1 - self.SyRec1_f.value) / self.ev1)*100,2)}% of policyholders left the group by end of simulation\n'
                            log4 = f'{round(sh_system.cell(2,21).value)} was the initial premium members were asked to pay.\n'
                            log5 = f'{self.SyRec19_f.value} is the final premium members were asked to pay.\n'
                            log6 = f'Premiums increased by {round((self.SyRec19_f.value / sh_system.cell(2,21).value)*100,2)}% by end of simulation\n'
                            log7 = f'self.SyRec 3 (period 0 finalize) = {sh_system.cell(3,5).value}\n'
                            log8 = f'{self.ev4*100}% of policyholders who were assigned to defect\n'
                            log9 = (sh_system.cell(3,5).value/self.ev1)*100
                            log9 = f'{round(log9,2)}% of policyholders who actually defected\n'
                            log10 = f'{(self.pv5)*100}% was the initial collapse threshold set for PV 5\n'
                        
                            L=[log1, log2, log3, log4, log5, log6, log7, log8, log9, log10]
                            for _log in L:
                                print(_log)
                            file1.writelines(L)
                            file1.close()

                            original_sys = os.path.join(root_sys, '1 System Database.xlsx')
                            original_us = os.path.join(root_sys, '2 User Database.xlsx')
                            edge_sys = os.path.join(root_sys, str(self.run) + '_' + '1 System Database.xlsx')
                            edge_us = os.path.join(root_sys, str(self.run) + '_' + '2 User Database.xlsx')

                            shutil.copyfile(original_sys, edge_sys)
                            shutil.copyfile(original_us, edge_us)
                            return
                    
                except Exception as e:
                    print(e)

            start_iter = start_iter+1
            if start_iter == 11:
            # if start_iter == 1:
                print(f'Iteration {start_iter} times complete! Please run the entire application again.')
                return

    def stopAction(self):
        sys.exit()
        print('Stopped')
        
    def clearAction(self):
        #clearing EV Variables
        self.entry1.delete(0,100)
        self.entry2.delete(0,100)
        self.entry3.delete(0,100)
        self.entry4.delete(0,100)
        self.entry5.delete(0,100)
        self.entry6.delete(0,100)
        self.entry7.delete(0,100)
        # clearing PV Variables
        self.entry1pv.delete(0,100)
        self.entry2pv.delete(0,100)
        self.entry3pv.delete(0,100)
        self.entry4pv.delete(0,100)
        self.entry5pv.delete(0,100)
        self.entry6pv.delete(0,100)

    def closeAction(self):
        self.app.destroy()
        
def main(matrix=False):
    if not matrix:
        simulator = Simulator(ui=True)
    elif matrix:
        generate_matrix()
        with open('matrix.json', 'r') as f:
            content = json.loads(f.read())
        matrix = content['matrix']
        simulator = Simulator(ui=False, _matrix=matrix)

if __name__ == '__main__':
    arg = sys.argv[-1]
    main_start = time.time()
    if sys.argv[-1] == '-m':
        main(matrix=True)
    else:
        main()
    print(f'Total program runtime: {(time.time()-main_start)/60} minutes')


# In[ ]:




